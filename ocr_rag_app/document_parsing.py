"""Document ingestion helpers, Office conversion, OCR, and file parsers.
文档入库辅助、Office 转换、OCR 和文件解析器。
"""

from .services import *


# =========================
# 文件处理函数
# =========================
def sanitize_uploaded_relative_name(file_name: str) -> str:
    raw_name = (file_name or "uploaded_file").replace("\\", "/")
    safe_parts = []
    invalid_chars = '<>:"|?*'
    translation = str.maketrans({char: "_" for char in invalid_chars})
    for part in raw_name.split("/"):
        cleaned = part.strip().translate(translation).rstrip(" .")
        if cleaned and cleaned not in {".", ".."}:
            safe_parts.append(cleaned)
    if not safe_parts:
        safe_parts = ["uploaded_file"]
    return os.path.join(*safe_parts)


def get_uploaded_relative_name(uploaded_file) -> str:
    return sanitize_uploaded_relative_name(getattr(uploaded_file, "name", "uploaded_file"))


def calculate_uploaded_file_sha256(uploaded_file) -> str:
    digest = hashlib.sha256()
    current_position = uploaded_file.tell() if hasattr(uploaded_file, "tell") else 0
    try:
        uploaded_file.seek(0)
        while True:
            chunk = uploaded_file.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    finally:
        try:
            uploaded_file.seek(current_position)
        except Exception:
            pass
    return digest.hexdigest()


def get_file_extension_from_name(file_name: str) -> str:
    return Path(file_name).suffix.lower().lstrip(".")


def is_legacy_office_file(file_name: str) -> bool:
    return get_file_extension_from_name(file_name) in LEGACY_OFFICE_TYPES


def is_supported_upload(file_name: str) -> bool:
    return get_file_extension_from_name(file_name) in SUPPORTED_TYPES


def is_soffice_executable(candidate: str) -> bool:
    return bool(candidate and os.path.exists(candidate) and not os.path.isdir(candidate))


def expand_soffice_candidates(path_value: str) -> List[str]:
    if not path_value:
        return []
    path_value = normalize_local_path(path_value, "")
    if os.path.isdir(path_value):
        return [
            os.path.join(path_value, "soffice"),
            os.path.join(path_value, "libreoffice"),
            os.path.join(path_value, "program", "soffice.exe"),
            os.path.join(path_value, "Contents", "MacOS", "soffice"),
        ]
    return [path_value]


def find_soffice_binary() -> Optional[str]:
    program_files = os.environ.get("PROGRAMFILES", r"C:\Program Files")
    program_files_x86 = os.environ.get("PROGRAMFILES(X86)", r"C:\Program Files (x86)")
    configured_candidates = expand_soffice_candidates(get_configured_soffice_path())
    for candidate in [
        *configured_candidates,
        shutil.which("soffice"),
        shutil.which("libreoffice"),
        "/Applications/LibreOffice.app/Contents/MacOS/soffice",
        "/usr/bin/soffice",
        "/usr/local/bin/soffice",
        "/usr/bin/libreoffice",
        "/usr/local/bin/libreoffice",
        "/snap/bin/libreoffice",
        os.path.join(program_files, "LibreOffice", "program", "soffice.exe"),
        os.path.join(program_files_x86, "LibreOffice", "program", "soffice.exe"),
    ]:
        if candidate and is_soffice_executable(candidate):
            return candidate
    return None


def find_textutil_binary() -> Optional[str]:
    return shutil.which("textutil")


def quote_command(command: List[str]) -> str:
    if platform.system() == "Windows":
        return subprocess.list2cmdline(command)
    return " ".join(shlex.quote(part) for part in command)


def linux_privileged_command(command: List[str]) -> List[str]:
    if platform.system() != "Linux":
        return command
    geteuid = getattr(os, "geteuid", None)
    if callable(geteuid) and geteuid() == 0:
        return command
    return ["sudo", "-n", *command]


def get_libreoffice_install_plan() -> Dict[str, Any]:
    system_name = platform.system()

    if system_name == "Darwin":
        brew = shutil.which("brew")
        if brew:
            command = [brew, "install", "--cask", "libreoffice"]
            return {
                "platform": "macOS",
                "commands": [command],
                "manual": quote_command(command),
            }
        return {
            "platform": "macOS",
            "commands": [],
            "manual": localized_text(
                "Install Homebrew first, then run: brew install --cask libreoffice",
                "先安装 Homebrew，然后执行：brew install --cask libreoffice",
                "先安裝 Homebrew，然後執行：brew install --cask libreoffice",
            ),
        }

    if system_name == "Windows":
        winget = shutil.which("winget")
        if winget:
            command = [
                winget,
                "install",
                "--id",
                "TheDocumentFoundation.LibreOffice",
                "-e",
                "--accept-source-agreements",
                "--accept-package-agreements",
            ]
            return {
                "platform": "Windows",
                "commands": [command],
                "manual": quote_command(command),
            }

        choco = shutil.which("choco")
        if choco:
            command = [choco, "install", "libreoffice-fresh", "-y"]
            return {
                "platform": "Windows",
                "commands": [command],
                "manual": quote_command(command),
            }

        return {
            "platform": "Windows",
            "commands": [],
            "manual": localized_text(
                "Install winget or Chocolatey, then run: winget install --id TheDocumentFoundation.LibreOffice -e",
                "安装 winget 或 Chocolatey 后执行：winget install --id TheDocumentFoundation.LibreOffice -e",
                "安裝 winget 或 Chocolatey 後執行：winget install --id TheDocumentFoundation.LibreOffice -e",
            ),
        }

    if system_name == "Linux":
        if shutil.which("apt-get"):
            commands = [
                linux_privileged_command(["apt-get", "update"]),
                linux_privileged_command(["apt-get", "install", "-y", "libreoffice"]),
            ]
            return {
                "platform": "Linux",
                "commands": commands,
                "manual": "sudo apt-get update && sudo apt-get install -y libreoffice",
            }

        if shutil.which("dnf"):
            command = linux_privileged_command(["dnf", "install", "-y", "libreoffice"])
            return {
                "platform": "Linux",
                "commands": [command],
                "manual": "sudo dnf install -y libreoffice",
            }

        if shutil.which("yum"):
            command = linux_privileged_command(["yum", "install", "-y", "libreoffice"])
            return {
                "platform": "Linux",
                "commands": [command],
                "manual": "sudo yum install -y libreoffice",
            }

        if shutil.which("zypper"):
            command = linux_privileged_command(["zypper", "--non-interactive", "install", "libreoffice"])
            return {
                "platform": "Linux",
                "commands": [command],
                "manual": "sudo zypper --non-interactive install libreoffice",
            }

        if shutil.which("pacman"):
            command = linux_privileged_command(["pacman", "-S", "--noconfirm", "libreoffice-still"])
            return {
                "platform": "Linux",
                "commands": [command],
                "manual": "sudo pacman -S --noconfirm libreoffice-still",
            }

        if shutil.which("snap"):
            command = linux_privileged_command(["snap", "install", "libreoffice"])
            return {
                "platform": "Linux",
                "commands": [command],
                "manual": "sudo snap install libreoffice",
            }

        return {
            "platform": "Linux",
            "commands": [],
            "manual": localized_text(
                "Install LibreOffice with your distribution package manager, for example: sudo apt-get install -y libreoffice",
                "请使用当前发行版包管理器安装 LibreOffice，例如：sudo apt-get install -y libreoffice",
                "請使用當前發行版套件管理器安裝 LibreOffice，例如：sudo apt-get install -y libreoffice",
            ),
        }

    return {
        "platform": system_name or localized_text("Unknown system", "未知系统", "未知系統"),
        "commands": [],
        "manual": localized_text(
            "No built-in automatic install plan is available for this system. Please install LibreOffice manually.",
            "当前系统未内置自动安装方案，请手动安装 LibreOffice。",
            "當前系統未內建自動安裝方案，請手動安裝 LibreOffice。",
        ),
    }


def install_libreoffice_automatically() -> Tuple[bool, str]:
    existing_binary = find_soffice_binary()
    if existing_binary:
        return True, localized_text("LibreOffice detected: ", "已检测到 LibreOffice：", "已檢測到 LibreOffice：") + existing_binary

    plan = get_libreoffice_install_plan()
    commands = plan.get("commands", [])
    manual_command = plan.get("manual", "")
    if not commands:
        return False, localized_text(
            "No available automatic install method found. Please run manually: ",
            "未找到可用的自动安装方式。请手动执行：",
            "未找到可用的自動安裝方式。請手動執行：",
        ) + manual_command

    logs = []
    for command in commands:
        result = run_subprocess(command, timeout=1800)
        command_text = quote_command(command)
        if result.returncode != 0:
            output = (result.stderr or result.stdout or "").strip()
            return (
                False,
                localized_text(
                    f"Automatic install failed.\nCommand: {command_text}\nReason: {output or 'Unknown error'}\nManual command: {manual_command}",
                    f"自动安装失败。\n命令：{command_text}\n原因：{output or '未知错误'}\n可手动执行：{manual_command}",
                    f"自動安裝失敗。\n命令：{command_text}\n原因：{output or '未知錯誤'}\n可手動執行：{manual_command}",
                ),
            )
        logs.append(localized_text("Completed: ", "完成：", "完成：") + command_text)

    installed_binary = find_soffice_binary()
    if installed_binary:
        return True, localized_text("LibreOffice installation completed: ", "LibreOffice 安装完成：", "LibreOffice 安裝完成：") + installed_binary

    return (
        False,
        localized_text(
            "The install command has run, but soffice is not detected in the current process. Restart the app and try again. If it still fails, run manually: ",
            "安装命令已执行，但当前进程还没有检测到 soffice。请重启应用后再试；如仍失败，可手动执行：",
            "安裝命令已執行，但當前進程尚未檢測到 soffice。請重啟應用後再試；如仍失敗，可手動執行：",
        )
        + manual_command,
    )


def legacy_conversion_target_ext(ext: str) -> str:
    return {
        "doc": "docx",
        "ppt": "pptx",
        "xls": "xlsx",
    }[ext]


def get_legacy_conversion_status(ext: str) -> Tuple[bool, str]:
    ext = ext.lower().lstrip(".")
    if ext not in LEGACY_OFFICE_TYPES:
        return True, localized_text("No conversion required", "无需转换", "無需轉換")

    if find_soffice_binary():
        return True, localized_text("Will convert with LibreOffice", "将使用 LibreOffice 转换", "將使用 LibreOffice 轉換")

    if ext == "doc" and find_textutil_binary():
        return True, localized_text("Will convert DOC with macOS textutil", "将使用 macOS textutil 转换 DOC", "將使用 macOS textutil 轉換 DOC")

    return False, localized_text(
        "LibreOffice is required to convert legacy .doc/.ppt/.xls files",
        "需要安装 LibreOffice 才能转换 .doc/.ppt/.xls 老格式",
        "需要安裝 LibreOffice 才能轉換 .doc/.ppt/.xls 舊格式",
    )


def is_processable_upload(file_name: str) -> bool:
    if not is_supported_upload(file_name):
        return False
    if not is_legacy_office_file(file_name):
        return True

    ext = get_file_extension_from_name(file_name)
    can_convert, _message = get_legacy_conversion_status(ext)
    return can_convert


def get_upload_support_status(file_name: str, auto_install_libreoffice: bool = True) -> Tuple[bool, str]:
    ext = get_file_extension_from_name(file_name)
    if not is_supported_upload(file_name):
        ext_label = ext or localized_text("No extension", "无扩展名", "無副檔名")
        return False, localized_text(
            f"Unsupported file type: {ext_label}",
            f"不支持的文件类型：{ext_label}",
            f"不支援的文件類型：{ext_label}",
        )
    if is_legacy_office_file(file_name):
        can_convert, message = get_legacy_conversion_status(ext)
        if not can_convert and auto_install_libreoffice:
            return True, localized_text(
                "Legacy Office file. LibreOffice will be installed automatically if possible before import.",
                "老版 Office 文件。导入前会尽量自动安装 LibreOffice 后再转换。",
                "舊版 Office 文件。導入前會盡量自動安裝 LibreOffice 後再轉換。",
            )
        return can_convert, message
    return True, localized_text("Supported", "支持", "支援")


def save_uploaded_file(uploaded_file, batch_id: Optional[str] = None) -> str:
    relative_name = get_uploaded_relative_name(uploaded_file)
    if batch_id:
        file_path = os.path.join(UPLOAD_DIR, batch_id, relative_name)
    else:
        safe_name = Path(relative_name).name
        file_path = os.path.join(UPLOAD_DIR, f"{uuid.uuid4().hex}_{safe_name}")
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    current_position = uploaded_file.tell() if hasattr(uploaded_file, "tell") else 0
    try:
        uploaded_file.seek(0)
        with open(file_path, "wb") as f:
            shutil.copyfileobj(uploaded_file, f, length=1024 * 1024)
    finally:
        try:
            uploaded_file.seek(current_position)
        except Exception:
            pass
    return file_path


def run_subprocess(command: List[str], timeout: int = 180) -> subprocess.CompletedProcess:
    return subprocess.run(
        command,
        check=False,
        capture_output=True,
        text=True,
        timeout=timeout,
    )


def convert_legacy_office_file(file_path: str) -> str:
    ext = get_file_extension(file_path)
    if ext not in LEGACY_OFFICE_TYPES:
        return file_path

    target_ext = legacy_conversion_target_ext(ext)
    output_dir = os.path.join(CONVERTED_DIR, uuid.uuid4().hex)
    os.makedirs(output_dir, exist_ok=True)
    converted_path = os.path.join(output_dir, f"{Path(file_path).stem}.{target_ext}")

    soffice_binary = find_soffice_binary()
    if soffice_binary:
        result = run_subprocess(
            [
                soffice_binary,
                "--headless",
                "--convert-to",
                target_ext,
                "--outdir",
                output_dir,
                file_path,
            ],
            timeout=240,
        )
        if result.returncode != 0:
            detail = (result.stderr or result.stdout or "").strip()
            raise RuntimeError(
                localized_text("LibreOffice conversion failed: ", "LibreOffice 转换失败：", "LibreOffice 轉換失敗：")
                + (detail or localized_text("Unknown error", "未知错误", "未知錯誤"))
            )

        if os.path.exists(converted_path):
            return converted_path

        converted_candidates = list(Path(output_dir).glob(f"*.{target_ext}"))
        if converted_candidates:
            return str(converted_candidates[0])

        detail = (result.stderr or result.stdout or "").strip()
        raise RuntimeError(
            localized_text("LibreOffice did not generate a ", "LibreOffice 未生成 ", "LibreOffice 未生成 ")
            + target_ext
            + localized_text(" file: ", " 文件：", " 文件：")
            + (detail or localized_text("No output", "无输出", "無輸出"))
        )

    if ext == "doc":
        textutil_binary = find_textutil_binary()
        if textutil_binary:
            result = run_subprocess(
                [
                    textutil_binary,
                    "-convert",
                    "docx",
                    "-output",
                    converted_path,
                    file_path,
                ],
                timeout=180,
            )
            if result.returncode == 0 and os.path.exists(converted_path):
                return converted_path
            detail = (result.stderr or result.stdout or "").strip()
            raise RuntimeError(
                localized_text("textutil DOC conversion failed: ", "textutil 转换 DOC 失败：", "textutil 轉換 DOC 失敗：")
                + (detail or localized_text("Unknown error", "未知错误", "未知錯誤"))
            )

    raise RuntimeError(
        localized_text(
            "LibreOffice is required to convert legacy .doc/.ppt/.xls files",
            "需要安装 LibreOffice 才能转换 .doc/.ppt/.xls 老格式文件",
            "需要安裝 LibreOffice 才能轉換 .doc/.ppt/.xls 舊格式文件",
        )
    )


def convert_office_file_to_pdf(file_path: str) -> str:
    output_dir = os.path.join(CONVERTED_DIR, uuid.uuid4().hex)
    os.makedirs(output_dir, exist_ok=True)
    soffice_binary = find_soffice_binary()
    if not soffice_binary:
        raise RuntimeError(
            localized_text(
                "LibreOffice is required to rasterize PPT/PPTX before OCR.",
                "需要安装 LibreOffice 才能将 PPT/PPTX 栅格化后 OCR。",
                "需要安裝 LibreOffice 才能將 PPT/PPTX 柵格化後 OCR。",
            )
        )

    result = run_subprocess(
        [
            soffice_binary,
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            output_dir,
            file_path,
        ],
        timeout=300,
    )
    if result.returncode != 0:
        detail = (result.stderr or result.stdout or "").strip()
        raise RuntimeError(
            localized_text("LibreOffice PDF conversion failed: ", "LibreOffice 转 PDF 失败：", "LibreOffice 轉 PDF 失敗：")
            + (detail or localized_text("Unknown error", "未知错误", "未知錯誤"))
        )

    expected_pdf = os.path.join(output_dir, f"{Path(file_path).stem}.pdf")
    if os.path.exists(expected_pdf):
        return expected_pdf

    pdf_candidates = list(Path(output_dir).glob("*.pdf"))
    if pdf_candidates:
        return str(pdf_candidates[0])

    detail = (result.stderr or result.stdout or "").strip()
    raise RuntimeError(
        localized_text("LibreOffice did not generate a PDF: ", "LibreOffice 未生成 PDF：", "LibreOffice 未生成 PDF：")
        + (detail or localized_text("No output", "无输出", "無輸出"))
    )


def get_file_extension(file_path: str) -> str:
    return Path(file_path).suffix.lower().lstrip(".")


class SpreadsheetRowLimitExceeded(RuntimeError):
    def __init__(self, row_count: int, row_limit: int):
        self.row_count = row_count
        self.row_limit = row_limit
        super().__init__(
            localized_text(
                f"Spreadsheet file skipped because it has {row_count} rows, exceeding the configured limit of {row_limit}.",
                f"表格文件行数 {row_count} 超过当前限制 {row_limit}，已跳过入库。",
                f"表格文件列數 {row_count} 超過目前限制 {row_limit}，已跳過入庫。",
            )
        )


def patch_openpyxl_data_validation_init() -> None:
    from openpyxl.worksheet.datavalidation import DataValidation
    import inspect

    if getattr(DataValidation, "_ocr_rag_unknown_kwargs_patch", False):
        return

    original_init = DataValidation.__init__
    accepted_parameters = set(inspect.signature(original_init).parameters.keys())

    def patched_init(self, *args, **kwargs):
        filtered_kwargs = {key: value for key, value in kwargs.items() if key in accepted_parameters}
        return original_init(self, *args, **filtered_kwargs)

    DataValidation.__init__ = patched_init
    DataValidation._ocr_rag_unknown_kwargs_patch = True


def load_workbook_safely(xlsx_path: str, **kwargs: Any):
    try:
        return load_workbook(xlsx_path, **kwargs)
    except TypeError as error:
        message = str(error)
        if "DataValidation.__init__()" in message and "unexpected keyword argument" in message:
            patch_openpyxl_data_validation_init()
            return load_workbook(xlsx_path, **kwargs)
        raise


def is_spreadsheet_file_name(file_name: str) -> bool:
    return get_file_extension_from_name(file_name) in {"xlsx", "xls", "csv"}


def count_xlsx_rows(xlsx_path: str) -> int:
    workbook = load_workbook_safely(xlsx_path, data_only=True, read_only=True)
    try:
        return sum(int(worksheet.max_row or 0) for worksheet in workbook.worksheets)
    finally:
        workbook.close()


def detect_text_file_encoding(file_path: str, sample_size: int = 256 * 1024) -> str:
    """Detect a practical text encoding for TXT/CSV files.
    为 TXT/CSV 文件检测可用文本编码。
    """
    with open(file_path, "rb") as source:
        sample = source.read(sample_size)
    if sample.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if sample.startswith((b"\xff\xfe", b"\xfe\xff")):
        return "utf-16"

    candidate_encodings = ["utf-8", "utf-8-sig", "gb18030", "big5", "utf-16", "cp1252"]
    for encoding in candidate_encodings:
        try:
            sample.decode(encoding)
            return encoding
        except UnicodeDecodeError:
            continue

    try:
        from charset_normalizer import from_bytes

        best_match = from_bytes(sample).best()
        if best_match and best_match.encoding:
            return best_match.encoding
    except Exception:
        pass
    return "utf-8"


def count_csv_rows(csv_path: str) -> int:
    encoding = detect_text_file_encoding(csv_path)
    with open(csv_path, "r", encoding=encoding, errors="replace", newline="") as source:
        return sum(1 for _line in source)


def prepare_spreadsheet_for_ingest(
    relative_name: str,
    file_path: str,
    skip_large_excel: bool,
    excel_row_limit: int,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> Tuple[str, Optional[int]]:
    if not is_spreadsheet_file_name(relative_name):
        return file_path, None

    prepared_path = file_path
    extension = get_file_extension_from_name(relative_name)
    if extension == "xls":
        if progress_callback:
            progress_callback(localized_text("Converting XLS for row-count check", "正在转换 XLS 以检查行数", "正在轉換 XLS 以檢查行數"))
        prepared_path = convert_legacy_office_file(file_path)

    if not skip_large_excel:
        return prepared_path, None

    if progress_callback:
        progress_callback(localized_text("Checking spreadsheet row count", "正在检查表格行数", "正在檢查表格列數"))
    row_count = count_csv_rows(prepared_path) if extension == "csv" else count_xlsx_rows(prepared_path)
    if row_count > excel_row_limit:
        raise SpreadsheetRowLimitExceeded(row_count=row_count, row_limit=excel_row_limit)
    return prepared_path, row_count


def get_source_type(file_path: str) -> str:
    ext = get_file_extension(file_path)
    if ext in ["png", "jpg", "jpeg", "webp", "bmp"]:
        return "image"
    return ext


def make_json_safe(value: Any) -> Any:
    if hasattr(value, "tolist"):
        return value.tolist()
    if isinstance(value, (list, tuple)):
        return [make_json_safe(item) for item in value]
    if isinstance(value, dict):
        return {str(key): make_json_safe(item) for key, item in value.items()}
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def get_indexed_ocr_geometry(page_result: Dict[str, Any], index: int) -> Any:
    for key in ["rec_boxes", "rec_polys", "dt_polys", "det_polys"]:
        values = page_result.get(key)
        if values is not None and index < len(values):
            return make_json_safe(values[index])
    return None


def ocr_single_image_with_boxes(image_path: str) -> Tuple[str, List[Dict[str, Any]]]:
    """
    Run OCR for one image file as-is.
    对单个图片文件按原样执行 OCR。
    """
    ocr_model = load_ocr_model()
    result = ocr_model.predict(image_path)
    lines = []
    boxes = []
    try:
        for page_result in result:
            if not page_result:
                continue

            rec_texts = page_result.get("rec_texts", [])
            rec_scores = page_result.get("rec_scores", [])
            for index, (text, score) in enumerate(zip(rec_texts, rec_scores)):
                if text and score >= 0.5:
                    lines.append(text)
                    boxes.append(
                        {
                            "text": text,
                            "score": float(score),
                            "box": get_indexed_ocr_geometry(page_result, index),
                        }
                    )
    finally:
        del result
        gc.collect()
    return "\n".join(lines), boxes


def get_image_dimensions(image_path: str) -> Tuple[int, int]:
    from PIL import Image

    Image.MAX_IMAGE_PIXELS = None
    with Image.open(image_path) as image:
        return image.size


def resolve_image_preprocess_settings(image_preprocess: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    config = dict(image_preprocess or {})
    mode = str(config.get("mode") or "auto")
    if mode not in IMAGE_PREPROCESS_PRESETS:
        mode = "auto"
    preset = dict(IMAGE_PREPROCESS_PRESETS[mode])
    if bool(config.get("custom", False)) and mode != "off":
        preset["max_side"] = max(0, int(config.get("max_side", preset["max_side"]) or 0))
        preset["max_pixels"] = max(0, int(config.get("max_pixels", preset["max_pixels"]) or 0))
        preset["jpeg_quality"] = min(100, max(60, int(config.get("jpeg_quality", preset["jpeg_quality"]) or 90)))
        preset["grayscale"] = bool(config.get("grayscale", preset["grayscale"]))
    preset["mode"] = mode
    return preset


def calculate_preprocessed_size(width: int, height: int, settings: Dict[str, Any]) -> Tuple[int, int]:
    max_side = int(settings.get("max_side", 0) or 0)
    max_pixels = int(settings.get("max_pixels", 0) or 0)
    if max_side <= 0 and max_pixels <= 0:
        return width, height

    scale = 1.0
    longest_side = max(width, height)
    if max_side > 0 and longest_side > max_side:
        scale = min(scale, max_side / longest_side)

    total_pixels = width * height
    if max_pixels > 0 and total_pixels > max_pixels:
        scale = min(scale, (max_pixels / total_pixels) ** 0.5)

    if scale >= 1.0:
        return width, height
    return max(1, int(width * scale)), max(1, int(height * scale))


def is_long_ocr_image(width: int, height: int) -> bool:
    shorter_side = max(1, min(width, height))
    return max(width, height) > OCR_TILE_MAX_SIDE_PIXELS and (max(width, height) / shorter_side) >= 3


def adapt_auto_image_preprocess_settings(settings: Dict[str, Any], width: int, height: int) -> Dict[str, Any]:
    settings = dict(settings)
    if settings.get("mode") != "auto":
        return settings

    total_pixels = width * height
    if is_long_ocr_image(width, height) or total_pixels >= 12_000_000:
        settings["max_side"] = min(int(settings.get("max_side", 2400) or 2400), 2200)
        settings["max_pixels"] = min(int(settings.get("max_pixels", 5_000_000) or 5_000_000), 4_000_000)
        settings["jpeg_quality"] = min(int(settings.get("jpeg_quality", 90) or 90), 90)
        settings["grayscale"] = True
    return settings


def preprocess_pil_image_for_ocr(image, settings: Dict[str, Any]):
    from PIL import Image

    width, height = image.size
    target_width, target_height = calculate_preprocessed_size(width, height, settings)
    if (target_width, target_height) != (width, height):
        resample_filter = getattr(getattr(Image, "Resampling", Image), "LANCZOS")
        image = image.resize((target_width, target_height), resample_filter)

    if bool(settings.get("grayscale", False)):
        return image.convert("L")

    if image.mode in {"RGBA", "LA", "P"}:
        background = Image.new("RGB", image.size, "white")
        if image.mode == "P":
            image = image.convert("RGBA")
        alpha = image.getchannel("A") if image.mode in {"RGBA", "LA"} else None
        background.paste(image.convert("RGB"), mask=alpha)
        return background
    return image.convert("RGB")


def save_pil_image_for_ocr(image, output_path: str, settings: Dict[str, Any]) -> None:
    quality = min(100, max(60, int(settings.get("jpeg_quality", 90) or 90)))
    image.save(output_path, format="JPEG", quality=quality, optimize=True)


def prepare_image_for_ocr(
    image_path: str,
    temp_dir: str,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> Tuple[str, bool]:
    settings = resolve_image_preprocess_settings(image_preprocess)
    if settings.get("mode") == "off":
        return image_path, False

    from PIL import Image

    Image.MAX_IMAGE_PIXELS = None
    with Image.open(image_path) as source_image:
        width, height = source_image.size
        settings = adapt_auto_image_preprocess_settings(settings, width, height)
        if is_long_ocr_image(width, height):
            return image_path, False
        target_size = calculate_preprocessed_size(width, height, settings)
        needs_color_change = bool(settings.get("grayscale", False)) or source_image.mode not in {"RGB", "L"}
        if target_size == (width, height) and not needs_color_change:
            return image_path, False

        prepared_image = preprocess_pil_image_for_ocr(source_image, settings)
        prepared_path = os.path.join(temp_dir, "ocr_preprocessed.jpg")
        try:
            save_pil_image_for_ocr(prepared_image, prepared_path, settings)
        finally:
            del prepared_image
            gc.collect()
    return prepared_path, True


def should_tile_image_for_ocr(image_path: str, image_preprocess: Optional[Dict[str, Any]] = None) -> bool:
    try:
        width, height = get_image_dimensions(image_path)
        if is_long_ocr_image(width, height):
            return True
        settings = resolve_image_preprocess_settings(image_preprocess)
        settings = adapt_auto_image_preprocess_settings(settings, width, height)
        if settings.get("mode") != "off":
            target_width, target_height = calculate_preprocessed_size(width, height, settings)
            width, height = target_width, target_height
        return max(width, height) > OCR_TILE_MAX_SIDE_PIXELS
    except Exception:
        return False


def iter_ocr_image_tiles(
    image_path: str,
    temp_dir: str,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> Iterator[Dict[str, Any]]:
    """
    Split very long or very large images into OCR-safe tiles one by one.
    将超长或超大图片逐块切成 PaddleOCR 更容易处理的小图，避免一次性持有整张 RGB 图。
    """
    from PIL import Image

    Image.MAX_IMAGE_PIXELS = None
    settings = resolve_image_preprocess_settings(image_preprocess)
    with Image.open(image_path) as source_image:
        width, height = source_image.size
        settings = adapt_auto_image_preprocess_settings(settings, width, height)
        if max(width, height) <= OCR_TILE_MAX_SIDE_PIXELS:
            tile_path = os.path.join(temp_dir, "tile_0001.jpg")
            tile_image = preprocess_pil_image_for_ocr(source_image, settings) if settings.get("mode") != "off" else source_image.convert("RGB")
            try:
                save_pil_image_for_ocr(tile_image, tile_path, settings)
            finally:
                del tile_image
                gc.collect()
            yield {
                "path": tile_path,
                "index": 1,
                "x0": 0,
                "y0": 0,
                "x1": width,
                "y1": height,
                "width": width,
                "height": height,
            }
            return

        step = max(1, OCR_TILE_MAX_SIDE_PIXELS - OCR_TILE_OVERLAP_PIXELS)
        tile_index = 0
        y0 = 0
        while y0 < height:
            y1 = min(height, y0 + OCR_TILE_MAX_SIDE_PIXELS)
            x0 = 0
            while x0 < width:
                x1 = min(width, x0 + OCR_TILE_MAX_SIDE_PIXELS)
                tile_index += 1
                tile_path = os.path.join(temp_dir, f"tile_{tile_index:04d}.jpg")
                tile = source_image.crop((x0, y0, x1, y1))
                tile = preprocess_pil_image_for_ocr(tile, settings) if settings.get("mode") != "off" else tile.convert("RGB")
                try:
                    save_pil_image_for_ocr(tile, tile_path, settings)
                finally:
                    del tile
                    gc.collect()
                yield {
                    "path": tile_path,
                    "index": tile_index,
                    "x0": x0,
                    "y0": y0,
                    "x1": x1,
                    "y1": y1,
                    "width": width,
                    "height": height,
                }
                if x1 >= width:
                    break
                x0 += step
            if y1 >= height:
                break
            y0 += step


def create_ocr_image_tiles(
    image_path: str,
    temp_dir: str,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    """
    Compatibility wrapper for older callers.
    兼容旧调用方；新流程使用 iter_ocr_image_tiles 逐块处理。
    """
    return list(iter_ocr_image_tiles(image_path, temp_dir, image_preprocess=image_preprocess))


def ocr_image_with_boxes(
    image_path: str,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> Tuple[str, List[Dict[str, Any]]]:
    """
    OCR an image and preserve source-location metadata.
    对图片做 OCR，并保留可用于后续定位原文的坐标信息。
    """
    text_parts = []
    all_boxes = []
    with tempfile.TemporaryDirectory(prefix="ocr_image_") as temp_dir:
        working_path, _is_temp = prepare_image_for_ocr(image_path, temp_dir, image_preprocess=image_preprocess)
        if not should_tile_image_for_ocr(working_path, image_preprocess=image_preprocess):
            return ocr_single_image_with_boxes(working_path)

        for tile in iter_ocr_image_tiles(working_path, temp_dir, image_preprocess=image_preprocess):
            try:
                tile_text, tile_boxes = ocr_single_image_with_boxes(tile["path"])
                if tile_text:
                    text_parts.append(tile_text)
                tile_meta = {
                    "tile_index": tile["index"],
                    "tile_x0": tile["x0"],
                    "tile_y0": tile["y0"],
                    "tile_x1": tile["x1"],
                    "tile_y1": tile["y1"],
                    "source_width": tile["width"],
                    "source_height": tile["height"],
                }
                for box in tile_boxes:
                    box["tile"] = tile_meta
                    all_boxes.append(box)
            finally:
                remove_file_quietly(tile.get("path", ""))
                gc.collect()
    return "\n".join(text_parts), all_boxes


def ocr_image(image_path: str, image_preprocess: Optional[Dict[str, Any]] = None) -> str:
    text, _boxes = ocr_image_with_boxes(image_path, image_preprocess=image_preprocess)
    return text


def extracted_image_output_path(source_file: str, image_name: str) -> str:
    image_ext = Path(image_name).suffix.lower()
    if image_ext not in OCR_IMAGE_EXTENSIONS:
        raise ValueError(localized_text("Unsupported image format: ", "不支持的图片格式：", "不支援的圖片格式：") + image_ext)

    safe_source = Path(source_file).stem[:40]
    return os.path.join(
        EXTRACTED_IMAGE_DIR,
        f"{safe_source}_{uuid.uuid4().hex}{image_ext}",
    )


def save_extracted_image(source_file: str, image_name: str, image_bytes: bytes) -> str:
    output_path = extracted_image_output_path(source_file, image_name)
    with open(output_path, "wb") as f:
        f.write(image_bytes)
    return output_path


def save_zip_member_to_extracted_image(archive: zipfile.ZipFile, source_file: str, media_name: str) -> str:
    output_path = extracted_image_output_path(source_file, media_name)
    with archive.open(media_name, "r") as src, open(output_path, "wb") as dst:
        shutil.copyfileobj(src, dst, length=1024 * 1024)
    return output_path


def serialize_ocr_boxes(ocr_boxes: List[Dict[str, Any]]) -> str:
    if not ocr_boxes:
        return ""
    total_count = len(ocr_boxes)
    limited_boxes = ocr_boxes[:OCR_BOX_METADATA_LIMIT]
    payload = {
        "items": limited_boxes,
        "total_count": total_count,
        "truncated": total_count > len(limited_boxes),
    }
    return json.dumps(payload, ensure_ascii=False)


def build_labeled_text(parts: List[Tuple[str, str]]) -> str:
    """
    合并直接解析文本和 OCR 文本，并去掉完全重复的行。
    """
    output = []
    seen_lines = set()

    for label, text in parts:
        lines = []
        for line in text.splitlines():
            normalized = line.strip()
            if not normalized:
                continue
            if normalized in seen_lines:
                continue
            seen_lines.add(normalized)
            lines.append(normalized)

        if lines:
            output.append(f"[{label}]\n" + "\n".join(lines))

    return "\n\n".join(output)


def format_indexed_progress(label: str, index: int, total: int) -> str:
    return localized_text(
        f"{label} ({index}/{total})",
        f"{label}（第 {index} 个，共 {total} 个）",
        f"{label}（第 {index} 個，共 {total} 個）",
    )


def iter_office_embedded_image_sections(
    file_path: str,
    source_type: str,
    media_prefix: str,
    image_preprocess: Optional[Dict[str, Any]] = None,
    progress_callback: Optional[Callable[[str], None]] = None,
    progress_label: str = "",
) -> Iterator[Dict[str, Any]]:
    """
    Stream embedded Office images through OCR one image at a time.
    逐张流式提取 Office 内嵌图片并 OCR，避免把所有图片或图片 bytes 一次性留在内存中。
    """
    try:
        with zipfile.ZipFile(file_path) as archive:
            media_names = [
                name
                for name in archive.namelist()
                if name.startswith(media_prefix) and Path(name).suffix.lower() in OCR_IMAGE_EXTENSIONS
            ]
            total_images = len(media_names)

            for image_index, media_name in enumerate(sorted(media_names), start=1):
                image_path = ""
                try:
                    if progress_callback and progress_label:
                        progress_callback(format_indexed_progress(progress_label, image_index, total_images))
                    image_path = save_zip_member_to_extracted_image(archive, file_path, media_name)
                    text, ocr_boxes = ocr_image_with_boxes(image_path, image_preprocess=image_preprocess)
                except Exception as e:
                    text = localized_text("OCR failed: ", "OCR 失败：", "OCR 失敗：") + str(e)
                    ocr_boxes = []

                ocr_boxes_json = serialize_ocr_boxes(ocr_boxes)
                del ocr_boxes
                section = make_section(
                    f"[{source_label('embedded_image_ocr')} {image_index}: {Path(media_name).name}]\n{text}",
                    {
                        "source_type": source_type,
                        "section_type": "embedded_image",
                        "image_index": image_index,
                        "image_name": Path(media_name).name,
                        "extract_method": "paddleocr_embedded_image",
                        "ocr_boxes": ocr_boxes_json,
                    },
                )
                if section:
                    yield section
                remove_file_quietly(image_path)
                del text, section
                release_memory_after_file()
    except zipfile.BadZipFile:
        pass


def extract_office_embedded_image_sections(
    file_path: str,
    source_type: str,
    media_prefix: str,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    return list(iter_office_embedded_image_sections(file_path, source_type, media_prefix, image_preprocess=image_preprocess))


def render_pdf_page_to_image(doc: fitz.Document, pdf_path: str, page_index: int, dpi: int = OCR_PDF_DPI) -> str:
    page = doc[page_index]
    base_scale = dpi / 72
    max_side = max(page.rect.width * base_scale, page.rect.height * base_scale)
    scale = base_scale
    if max_side > OCR_MAX_PAGE_SIDE_PIXELS:
        scale = OCR_MAX_PAGE_SIDE_PIXELS / max(page.rect.width, page.rect.height)

    matrix = fitz.Matrix(scale, scale)
    pix = page.get_pixmap(matrix=matrix, alpha=False, colorspace=fitz.csGRAY)
    image_path = os.path.join(
        UPLOAD_DIR,
        f"{Path(pdf_path).stem}_page_{page_index + 1}.png",
    )
    try:
        pix.save(image_path)
    finally:
        del pix
        gc.collect()
    return image_path


def remove_file_quietly(file_path: str) -> None:
    try:
        if file_path and os.path.exists(file_path):
            os.remove(file_path)
    except OSError:
        pass


def release_memory_after_file() -> None:
    gc.collect()
    try:
        torch = sys.modules.get("torch")
        if torch is None:
            return
        if hasattr(torch, "cuda") and torch.cuda.is_available():
            torch.cuda.empty_cache()
        if hasattr(torch, "mps") and hasattr(torch.mps, "empty_cache"):
            torch.mps.empty_cache()
    except Exception:
        pass


def make_section(text: str, metadata: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    text = text.strip()
    if not text:
        return None
    return {"text": text, "metadata": metadata}


def normalize_pdf_line(line: str) -> str:
    return " ".join((line or "").strip().split())


def is_probable_pdf_noise_line(text: str) -> bool:
    compact = normalize_pdf_line(text)
    if not compact:
        return True
    if len(compact) <= 6 and (
        compact.isdigit()
        or compact in {"目录", "目 录", "目錄"}
        or compact.strip("-—– ").isdigit()
    ):
        return True
    if re.match(r"^[\-.·•\s]+$", compact):
        return True
    if re.search(r"\.{4,}\s*\d+\s*$", compact):
        return True
    return False


def collect_pdf_layout_lines(page: fitz.Page) -> List[Dict[str, Any]]:
    page_dict = page.get_text("dict")
    lines = []
    for block in page_dict.get("blocks", []):
        if block.get("type") != 0:
            continue
        for line in block.get("lines", []):
            spans = line.get("spans", [])
            text = normalize_pdf_line("".join(span.get("text", "") for span in spans))
            if not text:
                continue
            sizes = [float(span.get("size", 0) or 0) for span in spans if span.get("text", "").strip()]
            bbox = line.get("bbox") or [0, 0, 0, 0]
            lines.append(
                {
                    "text": text,
                    "size": max(sizes) if sizes else 0.0,
                    "bbox": make_json_safe(bbox),
                    "y0": float(bbox[1]) if len(bbox) > 1 else 0.0,
                    "y1": float(bbox[3]) if len(bbox) > 3 else 0.0,
                    "page_height": float(page.rect.height),
                }
            )
    return lines


def find_repeated_pdf_margin_lines(layout_pages: List[List[Dict[str, Any]]]) -> set:
    page_count = len(layout_pages)
    if page_count < 3:
        return set()

    margin_counts = {}
    for lines in layout_pages:
        seen_on_page = set()
        for line in lines:
            page_height = line.get("page_height") or 1
            in_margin = line.get("y0", 0) < page_height * 0.10 or line.get("y1", 0) > page_height * 0.90
            text = normalize_pdf_line(line.get("text", ""))
            if in_margin and text and len(text) <= 80:
                seen_on_page.add(text)
        for text in seen_on_page:
            margin_counts[text] = margin_counts.get(text, 0) + 1

    threshold = max(3, int(page_count * 0.35))
    return {text for text, count in margin_counts.items() if count >= threshold}


def find_repeated_pdf_margin_lines_in_doc(doc: fitz.Document) -> set:
    """
    Detect repeated headers/footers without retaining all page layouts.
    不保留整份 PDF 的页面布局列表，只统计页眉页脚候选，降低大 PDF 内存峰值。
    """
    page_count = len(doc)
    if page_count < 3:
        return set()

    margin_counts = {}
    for page_index in range(page_count):
        lines = collect_pdf_layout_lines(doc[page_index])
        seen_on_page = set()
        for line in lines:
            page_height = line.get("page_height") or 1
            in_margin = line.get("y0", 0) < page_height * 0.10 or line.get("y1", 0) > page_height * 0.90
            text = normalize_pdf_line(line.get("text", ""))
            if in_margin and text and len(text) <= 80:
                seen_on_page.add(text)
        for text in seen_on_page:
            margin_counts[text] = margin_counts.get(text, 0) + 1
        del lines, seen_on_page
        gc.collect()

    threshold = max(3, int(page_count * 0.35))
    return {text for text, count in margin_counts.items() if count >= threshold}


def extract_pdf_table_text(page: fitz.Page) -> str:
    try:
        if not hasattr(page, "find_tables"):
            return ""
        tables = page.find_tables()
        table_parts = []
        for table_index, table in enumerate(tables, start=1):
            rows = table.extract()
            formatted_rows = []
            for row in rows:
                cells = [normalize_pdf_line(str(cell or "")) for cell in row]
                if any(cells):
                    formatted_rows.append(" | ".join(cells))
            if formatted_rows:
                table_parts.append(f"[{source_label('table')} {table_index}]\n" + "\n".join(formatted_rows))
        return "\n\n".join(table_parts)
    except Exception:
        return ""


def extract_pymupdf_layout_markdown(doc: fitz.Document, page_index: int) -> str:
    """
    Use PyMuPDF Layout when installed; fall back silently when unavailable.
    安装了 PyMuPDF Layout 时优先使用；不可用时静默回退。
    """
    try:
        import pymupdf.layout  # noqa: F401
        import pymupdf4llm

        markdown = pymupdf4llm.to_markdown(doc, pages=[page_index])
        if isinstance(markdown, list):
            markdown = "\n\n".join(str(item) for item in markdown)
        return str(markdown or "").strip()
    except Exception:
        return ""


def build_pdf_layout_text(
    page: fitz.Page,
    lines: List[Dict[str, Any]],
    repeated_margin_lines: set,
) -> str:
    sizes = [line.get("size", 0.0) for line in lines if line.get("size", 0.0) > 0]
    avg_size = sum(sizes) / len(sizes) if sizes else 0.0
    text_lines = []
    for line in lines:
        text = normalize_pdf_line(line.get("text", ""))
        if not text or text in repeated_margin_lines or is_probable_pdf_noise_line(text):
            continue
        is_title = avg_size > 0 and line.get("size", 0.0) >= avg_size * 1.25 and len(text) <= 100
        text_lines.append(f"{bracketed_label('title')} {text}" if is_title else text)

    table_text = extract_pdf_table_text(page)
    if table_text:
        text_lines.append(table_text)
    return "\n".join(text_lines).strip()


def iter_pdf_sections(
    pdf_path: str,
    ocr_threshold: int = 40,
    pdf_ocr_mode: str = "smart",
    progress_callback: Optional[Callable[[str], None]] = None,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> Iterator[Dict[str, Any]]:
    """
    PDF 混合解析：
    smart：优先提取 PDF 文字，只有文字很少的页才 OCR。
    force：每页都做整页 OCR，覆盖“文字 + 图片”混排，但内存和耗时最高。
    text：只提取 PDF 内置文字，不做 OCR。
    """
    pdf_ocr_mode = pdf_ocr_mode if pdf_ocr_mode in {"smart", "force", "text"} else "smart"
    doc = fitz.open(pdf_path)
    try:
        repeated_margin_lines = find_repeated_pdf_margin_lines_in_doc(doc)
        for page_index in range(len(doc)):
            page = doc[page_index]
            page_no = page_index + 1
            if progress_callback:
                progress_callback(
                    localized_text("Parsing PDF page ", "正在解析 PDF 第 ", "正在解析 PDF 第 ")
                    + f"{page_no}/{len(doc)}"
                    + localized_text("", " 页", " 頁")
                )
            layout_markdown = extract_pymupdf_layout_markdown(doc, page_index)
            page_lines = [] if layout_markdown else collect_pdf_layout_lines(page)
            direct_text = layout_markdown or build_pdf_layout_text(page, page_lines, repeated_margin_lines)
            del page_lines
            ocr_text = ""
            ocr_boxes = []
            extract_method = "pymupdf_layout" if layout_markdown else "pymupdf_text"
            compact_text_length = len(direct_text.replace("\n", "").strip())

            if pdf_ocr_mode == "force":
                image_path = render_pdf_page_to_image(doc, pdf_path, page_index)
                try:
                    if progress_callback:
                        progress_callback(
                            localized_text("OCR PDF page ", "正在 OCR PDF 第 ", "正在 OCR PDF 第 ")
                            + f"{page_no}/{len(doc)}"
                            + localized_text("", " 页", " 頁")
                        )
                    ocr_text, ocr_boxes = ocr_image_with_boxes(image_path, image_preprocess=image_preprocess)
                finally:
                    remove_file_quietly(image_path)
                text = build_labeled_text(
                    [
                        (source_label("direct_text"), direct_text),
                        (source_label("page_ocr_text"), ocr_text),
                    ]
                )
                extract_method = (
                    "pymupdf_layout+paddleocr_page"
                    if layout_markdown
                    else "pymupdf_text+paddleocr_page"
                )
            elif pdf_ocr_mode == "smart" and compact_text_length < ocr_threshold:
                image_path = render_pdf_page_to_image(doc, pdf_path, page_index)
                try:
                    if progress_callback:
                        progress_callback(
                            localized_text("OCR PDF page ", "正在 OCR PDF 第 ", "正在 OCR PDF 第 ")
                            + f"{page_no}/{len(doc)}"
                            + localized_text("", " 页", " 頁")
                        )
                    text, ocr_boxes = ocr_image_with_boxes(image_path, image_preprocess=image_preprocess)
                finally:
                    remove_file_quietly(image_path)
                extract_method = "paddleocr_fallback"
            else:
                text = direct_text

            ocr_boxes_json = serialize_ocr_boxes(ocr_boxes)
            del ocr_boxes
            section = make_section(
                f"[{source_label('page')} {page_no}]\n{text}",
                {
                    "source_type": "pdf",
                    "page": page_no,
                    "extract_method": extract_method,
                    "ocr_boxes": ocr_boxes_json,
                },
            )
            if section:
                yield section
            del layout_markdown, direct_text, ocr_text, text, section
            release_memory_after_file()
    finally:
        doc.close()


def extract_pdf_sections(
    pdf_path: str,
    ocr_threshold: int = 40,
    pdf_ocr_mode: str = "smart",
    progress_callback: Optional[Callable[[str], None]] = None,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    return list(
        iter_pdf_sections(
            pdf_path,
            ocr_threshold=ocr_threshold,
            pdf_ocr_mode=pdf_ocr_mode,
            progress_callback=progress_callback,
            image_preprocess=image_preprocess,
        )
    )


def iter_image_sections(
    image_path: str,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> Iterator[Dict[str, Any]]:
    text, ocr_boxes = ocr_image_with_boxes(image_path, image_preprocess=image_preprocess)
    section = make_section(
        text,
        {
            "source_type": "image",
            "extract_method": "paddleocr",
            "ocr_boxes": serialize_ocr_boxes(ocr_boxes),
        },
    )
    del ocr_boxes
    if section:
        yield section
    release_memory_after_file()


def extract_image_sections(
    image_path: str,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    return list(iter_image_sections(image_path, image_preprocess=image_preprocess))


def docx_row_cells(row) -> List[str]:
    return trim_empty_tail([cell.text.strip().replace("\n", " ") for cell in row.cells])


def build_docx_row_text(table_index: int, row_number: int, cells: List[str], header_cells: Optional[List[str]]) -> str:
    parts = [f"[{source_label('word_table')} {table_index}, {source_label('row')} {row_number}]"]
    if header_cells:
        max_column_count = max(len(cells), len(header_cells))
        for column_index in range(max_column_count):
            value = cells[column_index].strip() if column_index < len(cells) else ""
            if not value:
                continue
            header = header_cells[column_index].strip() if column_index < len(header_cells) else ""
            label = header or f"{source_label('column')} {column_index + 1}"
            parts.append(f"{label}: {value}")
    else:
        for column_index, value in enumerate(cells, start=1):
            if value.strip():
                parts.append(f"{source_label('column')} {column_index}: {value.strip()}")
    return "\n".join(parts)


def iter_docx_sections(
    docx_path: str,
    ocr_enhance: bool = True,
    progress_callback: Optional[Callable[[str], None]] = None,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> Iterator[Dict[str, Any]]:
    doc = Document(docx_path)
    try:
        if progress_callback:
            progress_callback(localized_text("Parsing Word paragraphs", "正在解析 Word 段落", "正在解析 Word 段落"))

        paragraph_parts = []
        for paragraph in doc.paragraphs:
            paragraph_text = paragraph.text.strip()
            if paragraph_text:
                paragraph_parts.append(paragraph_text)
            if len(paragraph_parts) >= 120:
                paragraph_section = make_section(
                    "\n".join(paragraph_parts),
                    {
                        "source_type": "docx",
                        "section_type": "paragraphs",
                        "extract_method": "python-docx",
                    },
                )
                if paragraph_section:
                    yield paragraph_section
                paragraph_parts.clear()

        paragraph_section = make_section(
            "\n".join(paragraph_parts),
            {
                "source_type": "docx",
                "section_type": "paragraphs",
                "extract_method": "python-docx",
            },
        )
        if paragraph_section:
            yield paragraph_section
        del paragraph_parts, paragraph_section

        for table_index, table in enumerate(doc.tables, start=1):
            if progress_callback:
                progress_callback(
                    localized_text("Parsing Word table ", "正在解析 Word 表格 ", "正在解析 Word 表格 ")
                    + f"{table_index}/{len(doc.tables)}"
                )
            header_cells = None
            for row_number, row in enumerate(table.rows, start=1):
                cells = docx_row_cells(row)
                if not any(cells):
                    continue
                if header_cells is None and looks_like_xlsx_header(cells):
                    header_cells = cells
                    header_section = make_section(
                        f"[{source_label('word_table')} {table_index}, {source_label('header_row')} {row_number}]\n"
                        + "\n".join(f"{source_label('column')} {index + 1}: {cell}" for index, cell in enumerate(cells) if cell),
                        {
                            "source_type": "docx",
                            "section_type": "table_header",
                            "table_index": table_index,
                            "row_number": row_number,
                            "row_range": str(row_number),
                            "extract_method": "python-docx",
                        },
                    )
                    if header_section:
                        yield header_section
                    continue

                row_section = make_section(
                    build_docx_row_text(table_index, row_number, cells, header_cells),
                    {
                        "source_type": "docx",
                        "section_type": "table_row",
                        "table_index": table_index,
                        "row_number": row_number,
                        "row_range": str(row_number),
                        "extract_method": "python-docx",
                    },
                )
                if row_section:
                    yield row_section

        if ocr_enhance:
            progress_label = localized_text("OCR for embedded Word images", "正在 OCR Word 内嵌图片", "正在 OCR Word 內嵌圖片")
            if progress_callback:
                progress_callback(progress_label)
            yield from iter_office_embedded_image_sections(
                file_path=docx_path,
                source_type="docx",
                media_prefix="word/media/",
                image_preprocess=image_preprocess,
                progress_callback=progress_callback,
                progress_label=progress_label,
            )
    finally:
        del doc
        release_memory_after_file()


def extract_docx_sections(
    docx_path: str,
    ocr_enhance: bool = True,
    progress_callback: Optional[Callable[[str], None]] = None,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    return list(
        iter_docx_sections(
            docx_path,
            ocr_enhance=ocr_enhance,
            progress_callback=progress_callback,
            image_preprocess=image_preprocess,
        )
    )


def shape_position_key(shape) -> Tuple[int, int]:
    return (
        int(getattr(shape, "top", 0) or 0),
        int(getattr(shape, "left", 0) or 0),
    )


def collect_ppt_shape_text(shape) -> List[str]:
    texts = []
    if getattr(shape, "has_text_frame", False):
        text = shape.text.strip()
        if text:
            texts.append(text)

    if getattr(shape, "has_table", False):
        rows = []
        for row in shape.table.rows:
            cells = [cell.text.strip().replace("\n", " ") for cell in row.cells]
            if any(cells):
                rows.append(" | ".join(cells))
        if rows:
            texts.append("\n".join(rows))

    if hasattr(shape, "shapes"):
        for child_shape in sorted(shape.shapes, key=shape_position_key):
            texts.extend(collect_ppt_shape_text(child_shape))

    return texts


def iter_pptx_sections(
    pptx_path: str,
    ocr_enhance: bool = True,
    progress_callback: Optional[Callable[[str], None]] = None,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> Iterator[Dict[str, Any]]:
    presentation = Presentation(pptx_path)
    try:
        for slide_index, slide in enumerate(presentation.slides, start=1):
            if progress_callback:
                progress_callback(
                    localized_text("Parsing PPT slide ", "正在解析 PPT 第 ", "正在解析 PPT 第 ")
                    + f"{slide_index}/{len(presentation.slides)}"
                    + localized_text("", " 页", " 頁")
                )
            slide_texts = []
            ordered_shapes = sorted(slide.shapes, key=shape_position_key)
            for shape in ordered_shapes:
                slide_texts.extend(collect_ppt_shape_text(shape))

            if slide_texts:
                first_text = slide_texts[0].strip()
                if first_text and len(first_text.replace("\n", "")) <= 100:
                    slide_texts[0] = f"{bracketed_label('title')}\n{first_text}"

            try:
                if slide.has_notes_slide:
                    notes = slide.notes_slide.notes_text_frame.text.strip()
                    if notes:
                        slide_texts.append(f"{source_label('note')}:\n" + notes)
            except Exception:
                pass

            section = make_section(
                f"[{source_label('slide')} {slide_index}]\n" + "\n\n".join(slide_texts),
                {
                    "source_type": "pptx",
                    "slide": slide_index,
                    "extract_method": "python-pptx",
                },
            )
            if section:
                yield section
            del slide_texts, ordered_shapes, section
            gc.collect()

        if ocr_enhance:
            progress_label = localized_text("OCR for embedded PPT images", "正在 OCR PPT 内嵌图片", "正在 OCR PPT 內嵌圖片")
            if progress_callback:
                progress_callback(progress_label)
            yield from iter_office_embedded_image_sections(
                file_path=pptx_path,
                source_type="pptx",
                media_prefix="ppt/media/",
                image_preprocess=image_preprocess,
                progress_callback=progress_callback,
                progress_label=progress_label,
            )
    finally:
        del presentation
        release_memory_after_file()


def extract_pptx_sections(
    pptx_path: str,
    ocr_enhance: bool = True,
    progress_callback: Optional[Callable[[str], None]] = None,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    return list(
        iter_pptx_sections(
            pptx_path,
            ocr_enhance=ocr_enhance,
            progress_callback=progress_callback,
            image_preprocess=image_preprocess,
        )
    )


def iter_presentation_visual_ocr_sections(
    presentation_path: str,
    progress_callback: Optional[Callable[[str], None]] = None,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> Iterator[Dict[str, Any]]:
    if progress_callback:
        progress_callback(
            localized_text(
                "Rasterizing PPT/PPTX to page images for OCR",
                "正在将 PPT/PPTX 栅格化为页面图像用于 OCR",
                "正在將 PPT/PPTX 柵格化為頁面圖像用於 OCR",
            )
    )
    pdf_path = convert_office_file_to_pdf(presentation_path)
    doc = fitz.open(pdf_path)
    try:
        for page_index in range(len(doc)):
            page_no = page_index + 1
            if progress_callback:
                progress_callback(
                    localized_text("OCR presentation page ", "正在 OCR 演示文稿第 ", "正在 OCR 簡報第 ")
                    + f"{page_no}/{len(doc)}"
                    + localized_text("", " 页", " 頁")
                )
            image_path = render_pdf_page_to_image(doc, pdf_path, page_index)
            try:
                text, ocr_boxes = ocr_image_with_boxes(image_path, image_preprocess=image_preprocess)
            finally:
                remove_file_quietly(image_path)

            section = make_section(
                f"[{source_label('slide')} {page_no}]\n{text}",
                {
                    "source_type": get_file_extension(presentation_path),
                    "page": page_no,
                    "slide": page_no,
                    "extract_method": "libreoffice_pdf+paddleocr_page",
                    "ocr_boxes": serialize_ocr_boxes(ocr_boxes),
                },
            )
            del ocr_boxes
            if section:
                yield section
            del text, section
            release_memory_after_file()
    finally:
        doc.close()
        remove_file_quietly(pdf_path)


def extract_presentation_visual_ocr_sections(
    presentation_path: str,
    progress_callback: Optional[Callable[[str], None]] = None,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    return list(
        iter_presentation_visual_ocr_sections(
            presentation_path,
            progress_callback=progress_callback,
            image_preprocess=image_preprocess,
        )
    )


def format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def trim_empty_tail(cells: List[str]) -> List[str]:
    cells = list(cells)
    while cells and not cells[-1]:
        cells.pop()
    return cells


def looks_like_xlsx_header(cells: List[str]) -> bool:
    non_empty_cells = [cell for cell in cells if cell.strip()]
    return len(non_empty_cells) >= 2


def build_xlsx_row_text(
    sheet_title: str,
    row_number: int,
    cells: List[str],
    header_cells: Optional[List[str]],
) -> str:
    parts = [f"[{source_label('worksheet')}: {sheet_title}, {source_label('row')} {row_number}]"]
    if header_cells:
        max_column_count = max(len(cells), len(header_cells))
        for column_index in range(max_column_count):
            value = cells[column_index].strip() if column_index < len(cells) else ""
            if not value:
                continue
            header = header_cells[column_index].strip() if column_index < len(header_cells) else ""
            label = header or f"{source_label('column')} {column_index + 1}"
            parts.append(f"{label}: {value}")
    else:
        for column_index, value in enumerate(cells, start=1):
            value = value.strip()
            if value:
                parts.append(f"{source_label('column')} {column_index}: {value}")
    return "\n".join(parts)


def iter_xlsx_sections(
    xlsx_path: str,
    rows_per_section: int = 80,
    ocr_enhance: bool = True,
    progress_callback: Optional[Callable[[str], None]] = None,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> Iterator[Dict[str, Any]]:
    workbook = load_workbook_safely(xlsx_path, data_only=True, read_only=True)

    try:
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            if progress_callback:
                progress_callback(
                    localized_text("Parsing Excel worksheet ", "正在解析 Excel 工作表 ", "正在解析 Excel 工作表 ")
                    + f"{sheet_index}/{len(workbook.worksheets)}: {worksheet.title}"
                )
            header_cells = None

            for row_number, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                cells = trim_empty_tail([format_cell_value(value).strip() for value in row])
                if not any(cells):
                    continue

                if header_cells is None and looks_like_xlsx_header(cells):
                    header_cells = cells
                    header_section = make_section(
                        f"[{source_label('worksheet')}: {worksheet.title}, {source_label('header_row')} {row_number}]\n"
                        + "\n".join(f"{source_label('column')} {index + 1}: {cell}" for index, cell in enumerate(cells) if cell),
                        {
                            "source_type": "xlsx",
                            "sheet": worksheet.title,
                            "row_number": row_number,
                            "row_range": str(row_number),
                            "section_type": "table_header",
                            "extract_method": "openpyxl",
                        },
                    )
                    if header_section:
                        yield header_section
                    continue

                section = make_section(
                    build_xlsx_row_text(worksheet.title, row_number, cells, header_cells),
                    {
                        "source_type": "xlsx",
                        "sheet": worksheet.title,
                        "row_number": row_number,
                        "row_range": str(row_number),
                        "section_type": "table_row",
                        "extract_method": "openpyxl",
                    },
                )
                if section:
                    yield section
                del cells, section
    finally:
        workbook.close()
        del workbook
        release_memory_after_file()

    if ocr_enhance:
        progress_label = localized_text("OCR for embedded Excel images", "正在 OCR Excel 内嵌图片", "正在 OCR Excel 內嵌圖片")
        if progress_callback:
            progress_callback(progress_label)
        yield from iter_office_embedded_image_sections(
            file_path=xlsx_path,
            source_type="xlsx",
            media_prefix="xl/media/",
            image_preprocess=image_preprocess,
            progress_callback=progress_callback,
            progress_label=progress_label,
        )


def extract_xlsx_sections(
    xlsx_path: str,
    rows_per_section: int = 80,
    ocr_enhance: bool = True,
    progress_callback: Optional[Callable[[str], None]] = None,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    return list(
        iter_xlsx_sections(
            xlsx_path,
            rows_per_section=rows_per_section,
            ocr_enhance=ocr_enhance,
            progress_callback=progress_callback,
            image_preprocess=image_preprocess,
        )
    )


def sniff_csv_dialect(sample: str):
    try:
        return csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except Exception:
        return csv.excel


def build_csv_row_text(
    row_number: int,
    cells: List[str],
    header_cells: Optional[List[str]],
) -> str:
    parts = [f"[{source_label('row')} {row_number}]"]
    if header_cells:
        max_column_count = max(len(cells), len(header_cells))
        for column_index in range(max_column_count):
            value = cells[column_index].strip() if column_index < len(cells) else ""
            if not value:
                continue
            header = header_cells[column_index].strip() if column_index < len(header_cells) else ""
            label = header or f"{source_label('column')} {column_index + 1}"
            parts.append(f"{label}: {value}")
    else:
        for column_index, value in enumerate(cells, start=1):
            value = value.strip()
            if value:
                parts.append(f"{source_label('column')} {column_index}: {value}")
    return "\n".join(parts)


def iter_csv_sections(
    csv_path: str,
    progress_callback: Optional[Callable[[str], None]] = None,
) -> Iterator[Dict[str, Any]]:
    encoding = detect_text_file_encoding(csv_path)
    if progress_callback:
        progress_callback(
            localized_text(
                f"Reading CSV file with detected encoding: {encoding}",
                f"正在读取 CSV 文件，检测到编码：{encoding}",
                f"正在讀取 CSV 文件，檢測到編碼：{encoding}",
            )
        )

    with open(csv_path, "r", encoding=encoding, errors="replace", newline="") as source:
        sample = source.read(8192)
        source.seek(0)
        reader = csv.reader(source, dialect=sniff_csv_dialect(sample))
        header_cells = None
        for row_number, row in enumerate(reader, start=1):
            cells = trim_empty_tail([format_cell_value(value).strip() for value in row])
            if not any(cells):
                continue

            if header_cells is None and looks_like_xlsx_header(cells):
                header_cells = cells
                header_section = make_section(
                    f"[{source_label('header_row')} {row_number}]\n"
                    + "\n".join(f"{source_label('column')} {index + 1}: {cell}" for index, cell in enumerate(cells) if cell),
                    {
                        "source_type": "csv",
                        "row_number": row_number,
                        "row_range": str(row_number),
                        "section_type": "table_header",
                        "extract_method": f"csv_decode:{encoding}",
                    },
                )
                if header_section:
                    yield header_section
                continue

            section = make_section(
                build_csv_row_text(row_number, cells, header_cells),
                {
                    "source_type": "csv",
                    "row_number": row_number,
                    "row_range": str(row_number),
                    "section_type": "table_row",
                    "extract_method": f"csv_decode:{encoding}",
                },
            )
            if section:
                yield section
            del cells, section
    release_memory_after_file()


def extract_csv_sections(csv_path: str, progress_callback: Optional[Callable[[str], None]] = None) -> List[Dict[str, Any]]:
    return list(iter_csv_sections(csv_path, progress_callback=progress_callback))


def iter_txt_sections(txt_path: str) -> Iterator[Dict[str, Any]]:
    encoding = detect_text_file_encoding(txt_path)
    buffer_lines = []
    buffer_size = 0
    section_index = 0
    max_section_chars = 20000

    def flush_buffer() -> Optional[Dict[str, Any]]:
        nonlocal buffer_lines, buffer_size, section_index
        text = "\n".join(buffer_lines).strip()
        buffer_lines = []
        buffer_size = 0
        if not text:
            return None
        section_index += 1
        return {
            "text": f"{source_label('direct_text')} {section_index}:\n{text}",
            "metadata": {
                "source_type": "txt",
                "source_label": source_label("direct_text"),
                "section_index": section_index,
                "extract_method": f"text_decode:{encoding}",
            },
        }

    try:
        with open(txt_path, "r", encoding=encoding, errors="replace", newline=None) as source:
            for line in source:
                normalized = line.rstrip("\n\r")
                buffer_lines.append(normalized)
                buffer_size += len(normalized) + 1
                if buffer_size >= max_section_chars:
                    section = flush_buffer()
                    if section:
                        yield section
                    release_memory_after_file()
        section = flush_buffer()
        if section:
            yield section
    except Exception as exc:
        raise ValueError(
            localized_text(
                f"Unable to decode the TXT file: {exc}",
                f"无法解码 TXT 文件：{exc}",
                f"無法解碼 TXT 文件：{exc}",
            )
        ) from exc
    finally:
        release_memory_after_file()


def extract_txt_sections(txt_path: str) -> List[Dict[str, Any]]:
    return list(iter_txt_sections(txt_path))


def iter_document_sections(
    file_path: str,
    ocr_enhance: bool = True,
    pdf_ocr_mode: str = "smart",
    ppt_visual_ocr: bool = True,
    progress_callback: Optional[Callable[[str], None]] = None,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> Iterator[Dict[str, Any]]:
    original_ext = get_file_extension(file_path)
    if original_ext in {"ppt", "pptx"} and ppt_visual_ocr:
        yield from iter_presentation_visual_ocr_sections(file_path, progress_callback=progress_callback, image_preprocess=image_preprocess)
        return

    file_path = convert_legacy_office_file(file_path)
    ext = get_file_extension(file_path)
    if ext == "pdf":
        yield from iter_pdf_sections(file_path, pdf_ocr_mode=pdf_ocr_mode, progress_callback=progress_callback, image_preprocess=image_preprocess)
        return
    if ext in ["png", "jpg", "jpeg", "webp", "bmp"]:
        if progress_callback:
            progress_callback(localized_text("OCR image", "正在 OCR 图片", "正在 OCR 圖片"))
        yield from iter_image_sections(file_path, image_preprocess=image_preprocess)
        return
    if ext == "docx":
        yield from iter_docx_sections(file_path, ocr_enhance=ocr_enhance, progress_callback=progress_callback, image_preprocess=image_preprocess)
        return
    if ext == "pptx":
        yield from iter_pptx_sections(file_path, ocr_enhance=ocr_enhance, progress_callback=progress_callback, image_preprocess=image_preprocess)
        return
    if ext == "xlsx":
        yield from iter_xlsx_sections(file_path, ocr_enhance=ocr_enhance, progress_callback=progress_callback, image_preprocess=image_preprocess)
        return
    if ext == "csv":
        yield from iter_csv_sections(file_path, progress_callback=progress_callback)
        return
    if ext == "txt":
        if progress_callback:
            progress_callback(localized_text("Reading TXT file", "正在读取 TXT 文件", "正在讀取 TXT 文件"))
        yield from iter_txt_sections(file_path)
        return
    raise ValueError(
        localized_text(
            "Supported formats are PDF, images, DOCX, PPTX, XLSX, CSV, and TXT. Legacy DOC, PPT, and XLS will be converted first when possible.",
            "当前支持 PDF、图片、DOCX、PPTX、XLSX、CSV、TXT；旧版 DOC、PPT、XLS 会先尝试转换为新版格式。",
            "目前支援 PDF、圖片、DOCX、PPTX、XLSX、CSV、TXT；舊版 DOC、PPT、XLS 會先嘗試轉換為新版格式。",
        )
    )


def extract_document_sections(
    file_path: str,
    ocr_enhance: bool = True,
    pdf_ocr_mode: str = "smart",
    ppt_visual_ocr: bool = True,
    progress_callback: Optional[Callable[[str], None]] = None,
    image_preprocess: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    return list(
        iter_document_sections(
            file_path,
            ocr_enhance=ocr_enhance,
            pdf_ocr_mode=pdf_ocr_mode,
            ppt_visual_ocr=ppt_visual_ocr,
            progress_callback=progress_callback,
            image_preprocess=image_preprocess,
        )
    )


def sections_to_text(sections: List[Dict[str, Any]]) -> str:
    return "\n\n".join(section["text"] for section in sections)


def sections_have_text(sections: List[Dict[str, Any]]) -> bool:
    return any(section.get("text", "").strip() for section in sections)
