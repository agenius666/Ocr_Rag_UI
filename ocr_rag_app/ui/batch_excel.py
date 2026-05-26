"""Batch Excel Q&A page with row-level JSON writeback.
批量 Excel 问答页面，支持按行 JSON 回写。
"""

from copy import copy

from openpyxl.utils import get_column_letter

from ..services import *
from .components import render_prompt_editor, render_search_results


def default_batch_output_mappings() -> List[Dict[str, Any]]:
    return [
        {
            "column": "D",
            "description": localized_text(
                "Supplier policy requirement",
                "供应商制度要求",
                "供應商制度要求",
            ),
        },
    ]


def css_content(value: str) -> str:
    return str(value or "").replace("\\", "\\\\").replace('"', '\\"')


def render_single_xlsx_uploader_css() -> None:
    upload_button_label = localized_text("Choose XLSX File", "选择 XLSX 文件", "選擇 XLSX 文件")
    st.markdown(
        f"""
        <style>
        div[data-testid="stFileUploader"]:has(input[accept*=".xlsx"]) section[data-testid="stFileUploaderDropzone"] button p {{
            display: none !important;
        }}
        div[data-testid="stFileUploader"]:has(input[accept*=".xlsx"]) section[data-testid="stFileUploaderDropzone"] button::after {{
            content: "{css_content(upload_button_label)}";
            font-size: 1rem;
            line-height: 1.5;
            margin-left: 0.45rem;
        }}
        div[data-testid="stFileUploader"]:has(input[accept*=".xlsx"]) div[data-testid="stFileUploaderFile"],
        div[data-testid="stFileUploader"]:has(input[accept*=".xlsx"]) section[data-testid="stFileUploaderDropzone"] div[data-testid="stFileUploaderFile"] {{
            display: flex !important;
        }}
        div[data-testid="stFileUploader"]:has(input[accept*=".xlsx"]) [data-testid="stFileUploaderFileLimit"],
        div[data-testid="stFileUploader"]:has(input[accept*=".xlsx"]) [data-testid="stFileUploaderDropzoneInstructions"],
        div[data-testid="stFileUploader"]:has(input[accept*=".xlsx"]) small {{
            display: none !important;
        }}
        </style>
        """,
        unsafe_allow_html=True,
    )


def normalize_excel_column(value: str) -> str:
    column = re.sub(r"[^A-Za-z]", "", str(value or "")).upper()
    return column or "A"


def cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_batch_output_mappings() -> List[Dict[str, Any]]:
    raw_value = get_config_value("batch_output_mappings_json", "")
    if raw_value:
        try:
            mappings = json.loads(raw_value)
            if isinstance(mappings, list) and mappings:
                return [
                    {
                        "column": normalize_excel_column(item.get("column", "D")),
                        "description": str(item.get("description", "") or "").strip(),
                    }
                    for item in mappings
                    if isinstance(item, dict)
                ] or default_batch_output_mappings()
        except Exception:
            pass
    return default_batch_output_mappings()


def save_batch_output_mappings(mappings: List[Dict[str, Any]]) -> None:
    set_config_value("batch_output_mappings_json", json.dumps(mappings, ensure_ascii=False))


def build_column_options(worksheet, header_row: int) -> List[str]:
    max_column = max(int(getattr(worksheet, "max_column", 1) or 1), 26)
    return [get_column_letter(index) for index in range(1, max_column + 1)]


def column_option_label(worksheet, header_row: int, column_letter: str) -> str:
    try:
        header_value = cell_to_text(worksheet[f"{column_letter}{header_row}"].value)
    except Exception:
        header_value = ""
    return f"{column_letter} - {header_value}" if header_value else column_letter


def build_row_values(worksheet, row_number: int, header_row: int) -> Dict[str, str]:
    values = {}
    for column_index in range(1, max(int(worksheet.max_column or 1), 26) + 1):
        column_letter = get_column_letter(column_index)
        cell_value = cell_to_text(worksheet[f"{column_letter}{row_number}"].value)
        values[column_letter] = cell_value
        header_name = cell_to_text(worksheet[f"{column_letter}{header_row}"].value)
        if header_name:
            values[header_name] = cell_value
    return values


def replace_excel_placeholders(template: str, row_values: Dict[str, str]) -> str:
    output = template or ""
    for key, value in row_values.items():
        output = output.replace("{" + key + "}", value)
    return output


def parse_row_numbers(row_spec: str, header_row: int, max_row: int) -> List[int]:
    first_data_row = min(max(header_row + 1, 1), max_row)
    if not str(row_spec or "").strip():
        return list(range(first_data_row, max_row + 1))

    selected_rows = []
    for part in re.split(r"[,，\s]+", str(row_spec).strip()):
        if not part:
            continue
        if "-" in part:
            left, right = part.split("-", 1)
            if left.strip().isdigit() and right.strip().isdigit():
                start = max(first_data_row, int(left))
                end = min(max_row, int(right))
                selected_rows.extend(range(min(start, end), max(start, end) + 1))
        elif part.isdigit():
            row_number = int(part)
            if first_data_row <= row_number <= max_row:
                selected_rows.append(row_number)
    return sorted(dict.fromkeys(selected_rows))


def build_json_template(mappings: List[Dict[str, Any]]) -> Dict[str, str]:
    template = {}
    for item in mappings:
        column = normalize_excel_column(item.get("column", ""))
        description = str(item.get("description", "") or column).strip()
        template["{" + column + "}"] = description
    return template


def extract_json_object(raw_text: str) -> Dict[str, Any]:
    text = str(raw_text or "").strip()
    text = re.sub(r"^```(?:json)?\s*", "", text, flags=re.I)
    text = re.sub(r"\s*```$", "", text)
    try:
        payload = json.loads(text)
        return payload if isinstance(payload, dict) else {}
    except Exception:
        pass

    match = re.search(r"\{.*\}", text, flags=re.S)
    if not match:
        return {}
    try:
        payload = json.loads(match.group(0))
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}


def copy_style_if_possible(source_cell, target_cell) -> None:
    try:
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.font = copy(source_cell.font)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
    except Exception:
        pass


def preserve_output_column_style(worksheet, target_cell, output_column: str, row_number: int, header_row: int) -> None:
    if getattr(target_cell, "has_style", False):
        return

    candidate_rows = []
    if row_number > header_row + 1:
        candidate_rows.append(row_number - 1)
    candidate_rows.append(header_row)

    for candidate_row in candidate_rows:
        try:
            source_cell = worksheet[f"{output_column}{candidate_row}"]
            if source_cell.has_style:
                copy_style_if_possible(source_cell, target_cell)
                return
        except Exception:
            continue


def should_skip_excel_row(worksheet, row_number: int, skip_column: str, min_chars: int) -> bool:
    if min_chars <= 0:
        return False
    value = cell_to_text(worksheet[f"{skip_column}{row_number}"].value)
    return len(value) < min_chars


def build_evidence_caption(search_results: List[Dict[str, Any]]) -> str:
    captions = []
    for index, item in enumerate(search_results, start=1):
        metadata = item.get("metadata") or {}
        captions.append(
            f"{source_label('material')} {index}: "
            f"{metadata.get('file_name', source_label('unknown_file'))} "
            f"{source_label('chunk_index')} {metadata.get('chunk_index', '')}"
        )
    return "；".join(captions)


def collect_batch_output_mappings_from_widgets(
    current_mappings: List[Dict[str, Any]],
    revision: int,
) -> List[Dict[str, Any]]:
    mappings = []
    for index, item in enumerate(current_mappings):
        mappings.append(
            {
                "column": normalize_excel_column(
                    st.session_state.get(f"batch_output_column_{revision}_{index}", item.get("column", "D"))
                ),
                "description": str(
                    st.session_state.get(f"batch_output_description_{revision}_{index}", item.get("description", ""))
                    or ""
                ).strip(),
            }
        )
    return mappings


def bump_batch_output_mapping_revision() -> None:
    st.session_state["batch_output_mapping_revision"] = int(
        st.session_state.get("batch_output_mapping_revision", 0)
    ) + 1


def remove_batch_output_mapping(index: int, current_mappings: List[Dict[str, Any]], revision: int) -> None:
    updated_mappings = collect_batch_output_mappings_from_widgets(current_mappings, revision)
    if len(updated_mappings) <= 1 or index >= len(updated_mappings):
        return
    updated_mappings.pop(index)
    st.session_state["batch_output_mappings"] = updated_mappings
    save_batch_output_mappings(updated_mappings)
    bump_batch_output_mapping_revision()


def add_batch_output_mapping(current_mappings: List[Dict[str, Any]], revision: int) -> None:
    updated_mappings = collect_batch_output_mappings_from_widgets(current_mappings, revision)
    updated_mappings.append(
        {
            "column": "D",
            "description": localized_text("Output field", "输出字段", "輸出欄位"),
        }
    )
    st.session_state["batch_output_mappings"] = updated_mappings
    save_batch_output_mappings(updated_mappings)
    bump_batch_output_mapping_revision()


def render_batch_output_mapping_editor(column_options: List[str], worksheet, header_row: int) -> List[Dict[str, Any]]:
    if "batch_output_mappings" not in st.session_state:
        st.session_state["batch_output_mappings"] = load_batch_output_mappings()
    revision = int(st.session_state.setdefault("batch_output_mapping_revision", 0))

    st.markdown(localized_text("#### Excel Writeback Configuration", "#### Excel 回写配置", "#### Excel 回寫配置"))
    st.caption(
        localized_text(
            "Choose output columns and describe what each column should contain. Length requirements can be written directly in the description.",
            "选择输出列，并填写每列要生成的内容说明；字数限制可直接写在输出说明里。",
            "選擇輸出欄，並填寫每欄要生成的內容說明；字數限制可直接寫在輸出說明裡。",
        )
    )

    current_mappings = st.session_state["batch_output_mappings"]
    normalized_mappings = []
    for index, item in enumerate(list(current_mappings)):
        row_cols = st.columns([1.1, 3.2, 0.4])
        current_column = normalize_excel_column(item.get("column", "D"))
        if current_column not in column_options:
            column_options.append(current_column)

        with row_cols[0]:
            output_column = st.selectbox(
                localized_text("Output Column", "输出列", "輸出欄"),
                column_options,
                index=column_options.index(current_column),
                format_func=lambda option: column_option_label(worksheet, header_row, option),
                key=f"batch_output_column_{revision}_{index}",
                help=localized_text(
                    "The Excel column to write this JSON field back to.",
                    "该 JSON 字段最终要写回的 Excel 列。",
                    "該 JSON 欄位最終要寫回的 Excel 欄。",
                ),
            )
        with row_cols[1]:
            description = st.text_input(
                localized_text("Output Description", "输出说明", "輸出說明"),
                value=str(item.get("description", "") or ""),
                key=f"batch_output_description_{revision}_{index}",
                placeholder=localized_text(
                    "Example: supplier policy requirement; no more than 150 words",
                    "例如：供应商制度要求；不超过150字",
                    "例如：供應商制度要求；不超過150字",
                ),
                help=localized_text(
                    "This becomes the instruction for the corresponding JSON field.",
                    "这会作为对应 JSON 字段的生成要求。",
                    "這會作為對應 JSON 欄位的生成要求。",
                ),
            )
        with row_cols[2]:
            st.write("")
            st.button(
                "-",
                key=f"remove_batch_output_mapping_{revision}_{index}",
                disabled=len(current_mappings) <= 1,
                on_click=remove_batch_output_mapping,
                args=(index, list(current_mappings), revision),
            )

        normalized_mappings.append(
            {
                "column": normalize_excel_column(output_column),
                "description": description.strip(),
            }
        )

    add_col, _ = st.columns([1, 5])
    with add_col:
        st.button(
            "+",
            key=f"add_batch_output_mapping_{revision}",
            on_click=add_batch_output_mapping,
            args=(list(current_mappings), revision),
        )

    st.session_state["batch_output_mappings"] = normalized_mappings
    save_batch_output_mappings(normalized_mappings)
    return normalized_mappings


def render_batch_excel_tab() -> None:
    st.subheader(localized_text("Batch Excel Analysis", "批量 Excel 分析", "批次 Excel 分析"))
    st.caption(
        localized_text(
            "Read an XLSX file row by row, retrieve evidence independently, ask the local LLM for strict JSON, and write results back to selected cells.",
            "按行读取 XLSX，逐行独立检索资料并调用大模型，要求模型返回严格 JSON 后写回指定单元格。",
            "按列讀取 XLSX，逐列獨立檢索資料並調用大模型，要求模型返回嚴格 JSON 後寫回指定儲存格。",
        )
    )

    render_single_xlsx_uploader_css()
    upload_revision = st.session_state.setdefault("batch_excel_upload_revision", 0)
    uploaded_file = st.file_uploader(
        localized_text("Upload One XLSX File", "上传一个 XLSX 文件", "上傳一個 XLSX 文件"),
        type=["xlsx"],
        accept_multiple_files=False,
        key=f"batch_excel_uploader_{upload_revision}",
        help=localized_text(
            "Only one .xlsx file is processed at a time. Use the clear button below to remove the current file.",
            "每次只处理一个 .xlsx 文件。可使用下方清除按钮移除当前文件。",
            "每次只處理一個 .xlsx 文件。可使用下方清除按鈕移除當前文件。",
        ),
    )
    if not uploaded_file:
        st.info(localized_text("Please upload one .xlsx file first.", "请先上传一个 .xlsx 文件。", "請先上傳一個 .xlsx 文件。"))
        return
    if st.button(
        localized_text("Clear Uploaded XLSX", "清除已上传 XLSX", "清除已上傳 XLSX"),
        key="clear_batch_excel_uploader",
        help=localized_text(
            "Clear the current uploaded workbook and choose another one.",
            "清除当前上传的工作簿，重新选择文件。",
            "清除當前上傳的活頁簿，重新選擇文件。",
        ),
    ):
        st.session_state["batch_excel_upload_revision"] = int(st.session_state.get("batch_excel_upload_revision", 0)) + 1
        st.rerun()

    file_bytes = uploaded_file.getvalue()
    try:
        preview_workbook = load_workbook_safely(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        st.error(localized_text(f"Failed to read Excel: {e}", f"读取 Excel 失败：{e}", f"讀取 Excel 失敗：{e}"))
        return

    sheet_names = preview_workbook.sheetnames
    selected_sheet = st.selectbox(
        localized_text("Worksheet", "选择工作表", "選擇工作表"),
        sheet_names,
        index=0,
        key="batch_sheet_name",
        help=localized_text(
            "The worksheet to read and write back.",
            "选择要读取并写回结果的工作表。",
            "選擇要讀取並寫回結果的工作表。",
        ),
    )
    preview_sheet = preview_workbook[selected_sheet]
    header_row = st.number_input(
        localized_text("Header Row", "表头所在行", "表頭所在列"),
        min_value=1,
        max_value=max(1, int(preview_sheet.max_row or 1)),
        value=min(max(1, get_int_config("batch_header_row", 1)), max(1, int(preview_sheet.max_row or 1))),
        step=1,
        key="batch_header_row_input",
        help=localized_text(
            "Column names are read from this row. You can reference either Excel letters like {B} or header names like {Supplier}.",
            "系统会从这一行读取列名；Prompt 中既可以用 {B}，也可以用 {供应商名称} 这类表头名。",
            "系統會從這一列讀取欄名；Prompt 中既可以用 {B}，也可以用 {供應商名稱} 這類表頭名。",
        ),
    )
    set_config_value("batch_header_row", int(header_row))

    st.info(
        localized_text(
            f"Current worksheet: {preview_sheet.max_row} rows and {preview_sheet.max_column} columns. Leave row range empty to process from the row after the header to the last row.",
            f"当前工作表共 {preview_sheet.max_row} 行，{preview_sheet.max_column} 列。空行号范围表示从表头下一行遍历到最后一行。",
            f"當前工作表共 {preview_sheet.max_row} 列，{preview_sheet.max_column} 欄。空列號範圍表示從表頭下一列遍歷到最後一列。",
        )
    )
    column_options = build_column_options(preview_sheet, int(header_row))

    default_prompt = get_config_value(
        "batch_prompt_template",
        localized_text(
            "Analyze supplier {B} against compliance requirement {C}, using the enterprise materials.",
            "请根据供应商【{B}】的合规要求【{C}】，结合企业资料进行分析。",
            "請根據供應商【{B}】的合規要求【{C}】，結合企業資料進行分析。",
        ),
    )
    prompt_template = st.text_area(
        localized_text("Row Prompt Template", "行级 Prompt 模板", "列級 Prompt 模板"),
        value=default_prompt,
        height=120,
        key="batch_prompt_template_input",
        placeholder=localized_text(
            "Example: Analyze supplier {B} against compliance requirement {C}, using the enterprise materials.",
            "例如：请根据供应商【{B}】的合规要求【{C}】，结合企业资料进行分析。",
            "例如：請根據供應商【{B}】的合規要求【{C}】，結合企業資料進行分析。",
        ),
        help=localized_text(
            "Write the reusable row-level task template. Placeholders such as {A}, {B}, {C}, or header names are replaced with current row values, and the final text becomes {row_prompt}.",
            "填写可复用的行级任务模板。{A}、{B}、{C} 或表头名会在每行调用前替换成该行实际单元格内容，替换后的最终文本就是 {row_prompt}。",
            "填寫可複用的列級任務模板。{A}、{B}、{C} 或表頭名會在每列調用前替換成該列實際儲存格內容，替換後的最終文字就是 {row_prompt}。",
        ),
    )
    set_config_value("batch_prompt_template", prompt_template)

    mappings = render_batch_output_mapping_editor(column_options, preview_sheet, int(header_row))
    json_template = build_json_template(mappings)
    output_columns = [normalize_excel_column(item.get("column", "")) for item in mappings]
    has_duplicate_output_columns = len(output_columns) != len(set(output_columns))
    if has_duplicate_output_columns:
        st.warning(
            localized_text(
                "Duplicate output columns are configured. Please adjust before starting batch analysis.",
                "输出列存在重复配置，请调整后再开始批量分析。",
                "輸出欄存在重複配置，請調整後再開始批次分析。",
            )
        )

    preview_payload = {
        "row_prompt_template": prompt_template,
        "json_template": json_template,
    }
    with st.expander(localized_text("Preview Prompt And JSON Template", "查看 Prompt 与 JSON 模板", "查看 Prompt 與 JSON 模板"), expanded=False):
        st.code(json.dumps(preview_payload, ensure_ascii=False, indent=2), language="json")

    placeholder_col = localized_text("Placeholder", "占位符", "佔位符")
    meaning_col = localized_text("Meaning", "含义", "含義")
    render_prompt_editor(
        "batch",
        default_batch_system_prompt_template(),
        default_batch_user_prompt_template(),
        localized_text(
            "Available placeholders: {row_prompt}, {context}, {json_template}, {json_template_text}, {language_instruction}",
            "可用占位符：{row_prompt}、{context}、{json_template}、{json_template_text}、{language_instruction}",
            "可用佔位符：{row_prompt}、{context}、{json_template}、{json_template_text}、{language_instruction}",
        ),
        [
            {
                placeholder_col: "{row_prompt}",
                meaning_col: localized_text(
                    "The final row prompt generated from the Row Prompt Template after replacing the current row's Excel placeholders.",
                    "由“行级 Prompt 模板”替换当前行 Excel 占位符后生成的最终行任务提示。",
                    "由「列級 Prompt 模板」替換當前列 Excel 佔位符後生成的最終列任務提示。",
                ),
            },
            {
                placeholder_col: "{context}",
                meaning_col: localized_text(
                    "Retrieved materials for the current row.",
                    "当前行召回并拼接后的检索资料。",
                    "當前列召回並拼接後的檢索資料。",
                ),
            },
            {
                placeholder_col: "{json_template}",
                meaning_col: localized_text(
                    "The JSON object schema generated from the writeback configuration.",
                    "根据 Excel 回写配置生成的 JSON 对象结构。",
                    "根據 Excel 回寫配置生成的 JSON 物件結構。",
                ),
            },
            {
                placeholder_col: "{json_template_text}",
                meaning_col: localized_text(
                    "Alias of {json_template}, kept for compatibility.",
                    "{json_template} 的兼容别名。",
                    "{json_template} 的相容別名。",
                ),
            },
            {
                placeholder_col: "{language_instruction}",
                meaning_col: localized_text(
                    "Output language instruction based on the current UI language.",
                    "根据当前界面语言生成的输出语言要求。",
                    "根據當前介面語言生成的輸出語言要求。",
                ),
            },
        ],
    )

    with st.expander(localized_text("Batch Processing And Retrieval Settings", "批量处理与检索设置", "批次處理與檢索設定"), expanded=False):
        row_col, skip_col, mode_col, scope_col = st.columns([1.2, 1.2, 1, 1])
        with row_col:
            row_spec = st.text_input(
                localized_text("Rows To Process", "处理行号 / 范围", "處理列號 / 範圍"),
                value=get_config_value("batch_row_spec", ""),
                key="batch_row_spec_input",
                placeholder=localized_text("Empty = all; e.g. 2-100 or 10", "空=全部；例如 2-100 或 10", "空=全部；例如 2-100 或 10"),
                help=localized_text(
                    "Leave empty to process every data row after the header. You can also enter one row, a range, or multiple values such as 2,5,9-20.",
                    "留空表示处理表头下一行到最后一行；也可以输入单行、范围或多个值，例如 2,5,9-20。",
                    "留空表示處理表頭下一列到最後一列；也可以輸入單列、範圍或多個值，例如 2,5,9-20。",
                ),
            )
            set_config_value("batch_row_spec", row_spec)
        with skip_col:
            skip_column = st.selectbox(
                localized_text("Skip Check Column", "跳过判断列", "跳過判斷欄"),
                column_options,
                index=column_options.index(get_config_value("batch_skip_column", "B"))
                if get_config_value("batch_skip_column", "B") in column_options
                else 0,
                format_func=lambda option: column_option_label(preview_sheet, int(header_row), option),
                key="batch_skip_column_input",
                help=localized_text(
                    "The app checks this column before processing a row. If the cell is too short, the row is skipped.",
                    "处理某一行前会检查这一列；如果单元格内容太短，该行会自动跳过。",
                    "處理某一列前會檢查這一欄；如果儲存格內容太短，該列會自動跳過。",
                ),
            )
            set_config_value("batch_skip_column", skip_column)
            skip_min_chars = st.number_input(
                localized_text("Skip If Shorter Than", "少于多少字符跳过", "少於多少字元跳過"),
                min_value=0,
                max_value=1000,
                value=max(0, get_int_config("batch_skip_min_chars", 10)),
                step=1,
                key="batch_skip_min_chars_input",
                help=localized_text(
                    "Default is 10. Set to 0 to disable row skipping by text length.",
                    "默认 10；设为 0 表示不按字符数跳过。",
                    "預設 10；設為 0 表示不按字元數跳過。",
                ),
            )
            set_config_value("batch_skip_min_chars", int(skip_min_chars))
        with mode_col:
            mode_options = list(LLM_MODE_OPTIONS.keys())
            saved_mode_label = get_config_value("batch_mode_label", "快速")
            if saved_mode_label not in mode_options:
                saved_mode_label = "快速"
            mode_label = st.radio(
                "回答模式",
                mode_options,
                index=mode_options.index(saved_mode_label),
                horizontal=True,
                key="batch_mode_label_input",
                help=localized_text(
                    "Fast and thinking modes use the model names and extra_body values configured in Settings.",
                    "快速 / 思考会使用配置中心里对应的模型名和 extra_body。",
                    "快速 / 思考會使用配置中心裡對應的模型名和 extra_body。",
                ),
            )
            set_config_value("batch_mode_label", mode_label)
            batch_mode = LLM_MODE_OPTIONS[mode_label]
        with scope_col:
            scope_options = ["全部资料", *DOC_CATEGORY_OPTIONS.keys()]
            saved_scope = get_config_value("batch_search_scope_label", "企业资料")
            if saved_scope not in scope_options:
                saved_scope = "企业资料"
            search_scope_label = st.selectbox(
                localized_text("Search Scope", "检索范围", "檢索範圍"),
                scope_options,
                index=scope_options.index(saved_scope),
                key="batch_search_scope_label_input",
                help=localized_text(
                    "Choose which ingested materials are searched for each Excel row.",
                    "选择每一行要检索哪一类已入库资料。",
                    "選擇每一列要檢索哪一類已入庫資料。",
                ),
            )
            set_config_value("batch_search_scope_label", search_scope_label)
            search_category = None if search_scope_label == "全部资料" else DOC_CATEGORY_OPTIONS[search_scope_label]

        retrieval_col1, retrieval_col2, retrieval_col3, retrieval_col4 = st.columns(4)
        with retrieval_col1:
            top_k = st.slider(
                localized_text("Top-K Chunks", "召回片段数量", "召回片段數量"),
                min_value=1,
                max_value=20,
                value=get_int_config("batch_top_k", DEFAULT_RAG_TOP_K),
                step=1,
                key="batch_top_k_input",
                help=localized_text(
                    "How many final chunks are sent to the LLM for each row. Smaller values are cleaner; larger values provide more evidence.",
                    "每一行最终送入大模型的片段数量。数值小更干净，数值大证据更多但噪音也更多。",
                    "每一列最終送入大模型的片段數量。數值小更乾淨，數值大證據更多但噪音也更多。",
                ),
            )
            set_config_value("batch_top_k", top_k)
        with retrieval_col2:
            use_distance_threshold = st.checkbox(
                localized_text("Enable Distance Threshold", "启用距离阈值", "啟用距離閾值"),
                value=get_bool_config("batch_use_distance_threshold", True),
                key="batch_use_distance_threshold_input",
                help=localized_text(
                    "Filter weak matches by vector distance. Disable it only when recall is too sparse.",
                    "按向量距离过滤弱相关片段；只有召回太少时才建议关闭。",
                    "按向量距離過濾弱相關片段；只有召回太少時才建議關閉。",
                ),
            )
            set_bool_config("batch_use_distance_threshold", use_distance_threshold)
            max_distance = st.slider(
                localized_text("Max Distance", "最大距离", "最大距離"),
                min_value=0.20,
                max_value=2.00,
                value=min(max(get_float_config("batch_max_distance", DEFAULT_VECTOR_MAX_DISTANCE), 0.20), 2.00),
                step=0.05,
                key="batch_max_distance_input",
                disabled=not use_distance_threshold,
                help=localized_text(
                    "Lower values are stricter. If useful evidence is filtered out, increase this value gradually.",
                    "数值越小越严格；如果有用资料被过滤，可逐步调大。",
                    "數值越小越嚴格；如果有用資料被過濾，可逐步調大。",
                ),
            )
            set_config_value("batch_max_distance", max_distance)
        with retrieval_col3:
            use_hybrid = st.checkbox(
                localized_text("Enable Hybrid Search", "启用混合检索", "啟用混合檢索"),
                value=get_bool_config("batch_use_hybrid", DEFAULT_USE_HYBRID_SEARCH),
                key="batch_use_hybrid_input",
                help=localized_text(
                    "Combine vector semantic search with keyword search. Useful for supplier names, clause names, and exact terms.",
                    "结合向量语义检索和关键词检索，适合供应商名称、条款名称、精确术语。",
                    "結合向量語義檢索和關鍵詞檢索，適合供應商名稱、條款名稱、精確術語。",
                ),
            )
            set_bool_config("batch_use_hybrid", use_hybrid)
            use_reranker = st.checkbox(
                localized_text("Enable Reranker", "启用重排模型", "啟用重排模型"),
                value=get_bool_config("batch_use_reranker", DEFAULT_USE_RERANKER),
                key="batch_use_reranker_input",
                help=localized_text(
                    "Retrieve more candidates first, then rerank them. This can improve accuracy but costs more memory and time.",
                    "先召回更多候选片段，再用重排模型排序。准确率可能更高，但会增加内存和耗时。",
                    "先召回更多候選片段，再用重排模型排序。準確率可能更高，但會增加記憶體和耗時。",
                ),
            )
            set_bool_config("batch_use_reranker", use_reranker)
        with retrieval_col4:
            fetch_k = st.slider(
                localized_text("Candidate Count", "候选召回数", "候選召回數"),
                min_value=top_k,
                max_value=50,
                value=min(max(get_int_config("batch_fetch_k", DEFAULT_RETRIEVAL_FETCH_K), top_k), 50),
                step=1,
                key="batch_fetch_k_input",
                help=localized_text(
                    "Candidate pool size for hybrid search and reranking. It must be at least Top-K.",
                    "混合检索和重排使用的候选池大小，必须不小于召回片段数量。",
                    "混合檢索和重排使用的候選池大小，必須不小於召回片段數量。",
                ),
            )
            set_config_value("batch_fetch_k", fetch_k)

    row_numbers = parse_row_numbers(row_spec, int(header_row), int(preview_sheet.max_row or 1))
    st.caption(
        localized_text(
            f"Estimated rows to process: {len(row_numbers)}. Each row is retrieved and answered independently without multi-turn context.",
            f"预计处理 {len(row_numbers)} 行；每行独立检索、独立调用模型，不带入多轮上下文。",
            f"預計處理 {len(row_numbers)} 列；每列獨立檢索、獨立調用模型，不帶入多輪上下文。",
        )
    )
    large_batch_threshold = 200
    confirm_large_batch = True
    if len(row_numbers) > large_batch_threshold:
        st.warning(
            localized_text(
                f"The current run will process {len(row_numbers)} rows and call the local LLM row by row. This may take a long time. Try a small row range first.",
                f"当前预计处理 {len(row_numbers)} 行，会逐行调用本地大模型，耗时可能很长。建议先用行号范围小批量验证。",
                f"當前預計處理 {len(row_numbers)} 列，會逐列調用本地大模型，耗時可能很長。建議先用列號範圍小批量驗證。",
            )
        )
        confirm_large_batch = st.checkbox(
            localized_text(
                f"Confirm processing more than {large_batch_threshold} rows",
                f"确认处理超过 {large_batch_threshold} 行",
                f"確認處理超過 {large_batch_threshold} 列",
            ),
            value=False,
            key="batch_confirm_large_run",
        )

    if st.button(
        localized_text("Start Batch Analysis And Generate Excel", "开始批量分析并生成新 Excel", "開始批次分析並生成新 Excel"),
        type="primary",
        key="start_batch_excel_analysis",
    ):
        if not mappings:
            st.warning(localized_text("Please configure at least one output column.", "请至少配置一个输出列。", "請至少配置一個輸出欄。"))
            st.stop()
        if has_duplicate_output_columns:
            st.warning(
                localized_text(
                    "Duplicate output columns are configured. Please adjust before starting batch analysis.",
                    "输出列存在重复配置，请调整后再开始批量分析。",
                    "輸出欄存在重複配置，請調整後再開始批次分析。",
                )
            )
            st.stop()
        if not row_numbers:
            st.warning(localized_text("No rows to process. Please check the row range.", "没有可处理的行，请检查行号范围。", "沒有可處理的列，請檢查列號範圍。"))
            st.stop()
        if len(row_numbers) > large_batch_threshold and not confirm_large_batch:
            st.warning(
                localized_text(
                    "Please confirm the large batch run first, or reduce the row range.",
                    "请先勾选大批量处理确认，或缩小处理行号范围。",
                    "請先勾選大批次處理確認，或縮小處理列號範圍。",
                )
            )
            st.stop()

        output_workbook = load_workbook_safely(io.BytesIO(file_bytes), data_only=False, read_only=False)
        worksheet = output_workbook[selected_sheet]
        progress_bar = st.progress(
            0,
            text=localized_text("Preparing batch analysis...", "准备开始批量分析...", "準備開始批次分析..."),
        )
        status_box = st.empty()
        result_rows = []
        last_search_results = []
        row_key = localized_text("Row", "行号", "列號")
        status_key = localized_text("Status", "状态", "狀態")
        message_key = localized_text("Message", "说明", "說明")

        success_count = 0
        skipped_count = 0
        failed_count = 0

        for row_index, row_number in enumerate(row_numbers, start=1):
            if should_skip_excel_row(worksheet, row_number, skip_column, int(skip_min_chars)):
                skipped_count += 1
                result_rows.append(
                    {
                        row_key: row_number,
                        status_key: localized_text("Skipped", "跳过", "跳過"),
                        message_key: localized_text(
                            f"{skip_column}{row_number} is shorter than the minimum length",
                            f"{skip_column}{row_number} 字符数不足",
                            f"{skip_column}{row_number} 字元數不足",
                        ),
                    }
                )
                progress_bar.progress(
                    row_index / len(row_numbers),
                    text=localized_text(f"Skipped row {row_number}", f"已跳过第 {row_number} 行", f"已跳過第 {row_number} 列"),
                )
                continue

            row_values = build_row_values(worksheet, row_number, int(header_row))
            row_prompt = replace_excel_placeholders(prompt_template, row_values)
            retrieval_query = row_prompt.strip()
            status_box.info(
                localized_text(
                    f"Processing row {row_number}: retrieving evidence and calling the model...",
                    f"正在处理第 {row_number} 行：检索资料并调用模型...",
                    f"正在處理第 {row_number} 列：檢索資料並調用模型...",
                )
            )

            try:
                search_results = search_vector_store_multi_query(
                    [retrieval_query],
                    top_k=top_k,
                    doc_category=search_category,
                    max_distance=max_distance if use_distance_threshold else None,
                    fetch_k=fetch_k,
                    use_hybrid=use_hybrid,
                    use_reranker=use_reranker,
                )
                last_search_results = search_results
                raw_answer = ask_llm_batch_excel(
                    row_prompt=row_prompt,
                    json_template=json_template,
                    search_results=search_results,
                    mode=batch_mode,
                )
                parsed_answer = extract_json_object(raw_answer)
                if not parsed_answer:
                    raise ValueError(localized_text("The model did not return parseable JSON", "模型未返回可解析的 JSON", "模型未返回可解析的 JSON"))

                for mapping in mappings:
                    output_column = normalize_excel_column(mapping["column"])
                    key_with_braces = "{" + output_column + "}"
                    value = parsed_answer.get(key_with_braces, parsed_answer.get(output_column, ""))
                    if not value and "附件" in mapping.get("description", ""):
                        value = build_evidence_caption(search_results)
                    target_cell = worksheet[f"{output_column}{row_number}"]
                    preserve_output_column_style(
                        worksheet,
                        target_cell,
                        output_column,
                        row_number,
                        int(header_row),
                    )
                    target_cell.value = cell_to_text(value)

                success_count += 1
                result_rows.append(
                    {
                        row_key: row_number,
                        status_key: localized_text("Succeeded", "成功", "成功"),
                        message_key: localized_text(
                            f"Wrote back {len(mappings)} cells",
                            f"写回 {len(mappings)} 个单元格",
                            f"寫回 {len(mappings)} 個儲存格",
                        ),
                    }
                )
                progress_bar.progress(
                    row_index / len(row_numbers),
                    text=localized_text(f"Completed row {row_number}", f"已完成第 {row_number} 行", f"已完成第 {row_number} 列"),
                )
            except Exception as e:
                failed_count += 1
                result_rows.append({row_key: row_number, status_key: localized_text("Failed", "失败", "失敗"), message_key: str(e)})
                progress_bar.progress(
                    row_index / len(row_numbers),
                    text=localized_text(f"Row {row_number} failed", f"第 {row_number} 行处理失败", f"第 {row_number} 列處理失敗"),
                )

        output = io.BytesIO()
        output_workbook.save(output)
        output_workbook.close()
        output.seek(0)
        status_box.success(
            localized_text(
                f"Batch analysis completed: {success_count} succeeded, {skipped_count} skipped, {failed_count} failed.",
                f"批量分析完成：成功 {success_count} 行，跳过 {skipped_count} 行，失败 {failed_count} 行。",
                f"批次分析完成：成功 {success_count} 列，跳過 {skipped_count} 列，失敗 {failed_count} 列。",
            )
        )
        st.download_button(
            localized_text("Download Filled Excel", "下载回写后的 Excel", "下載回寫後的 Excel"),
            data=output.getvalue(),
            file_name=f"batch_analysis_{Path(uploaded_file.name).stem}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="download_batch_excel_result",
        )
        st.dataframe(result_rows, width="stretch", hide_index=True)
        if last_search_results:
            with st.expander(localized_text("View Retrieved Materials For Last Row", "查看最后一行检索资料", "查看最後一列檢索資料"), expanded=False):
                render_search_results(last_search_results)
